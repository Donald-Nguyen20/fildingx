"""
ui/notebooklm_window.py — NotebookLM integration window.
"""
import asyncio
import subprocess
import sys
import os
import threading
from pathlib import Path
from datetime import datetime


def _load_source_map() -> dict:
    """Load nlm_source_map.json. Returns {} nếu chưa có hoặc lỗi.
    Hỗ trợ cả format cũ {"name": "path"} và mới {"name": {"path":..., "notebooks":[...]}}.
    """
    import json
    try:
        from paths import NLM_SOURCE_MAP
        with open(NLM_SOURCE_MAP, "r", encoding="utf-8") as f:
            raw = json.load(f)
        # Migrate format cũ
        migrated = {}
        for k, v in raw.items():
            if isinstance(v, str):
                migrated[k] = {"path": v, "notebooks": []}
            else:
                migrated[k] = v
        return migrated
    except Exception:
        return {}


def _save_source_map(src_map: dict):
    """Ghi nlm_source_map.json, bỏ các entry file không còn tồn tại."""
    import json
    try:
        from paths import NLM_SOURCE_MAP
        # Loại entry path không hợp lệ
        clean = {k: v for k, v in src_map.items()
                 if isinstance(v, dict) and os.path.isfile(v.get("path", ""))}
        with open(NLM_SOURCE_MAP, "w", encoding="utf-8") as f:
            json.dump(clean, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _upsert_source_map(source_title: str, file_path: str, notebook_id: str = ""):
    """Thêm hoặc cập nhật entry trong source map."""
    src_map = _load_source_map()
    entry = src_map.get(source_title, {"path": file_path, "notebooks": []})
    entry["path"] = file_path
    if notebook_id and notebook_id not in entry.get("notebooks", []):
        entry.setdefault("notebooks", []).append(notebook_id)
    src_map[source_title] = entry
    _save_source_map(src_map)


def _lookup_source_path(source_title: str) -> str | None:
    """Tìm đường dẫn local từ tên source. Trả None nếu không có."""
    src_map = _load_source_map()
    t = source_title.lower().strip()

    def _try(key: str) -> dict | None:
        return src_map.get(key)

    # 1. Exact match
    entry = _try(source_title)
    # 2. Case-insensitive
    if not entry:
        entry = next((v for k, v in src_map.items() if k.lower() == t), None)
    # 3. NbLM bỏ extension → thêm lại .pdf rồi thử
    if not entry:
        entry = _try(source_title + ".pdf") or next(
            (v for k, v in src_map.items() if k.lower() == t + ".pdf"), None
        )
    # 4. Key có extension nhưng title không có → bỏ extension của key
    if not entry:
        entry = next(
            (v for k, v in src_map.items()
             if os.path.splitext(k)[0].lower() == t), None
        )
    # 5. Partial match — title là phần đầu của key (NbLM cắt ngắn)
    if not entry:
        entry = next(
            (v for k, v in src_map.items()
             if k.lower().startswith(t) or t.startswith(k.lower().rstrip(".pdf").rstrip())), None
        )
    if not entry:
        return None
    path = entry.get("path", "")
    return path if os.path.isfile(path) else None


def is_nlm_logged_in() -> bool:
    """Kiểm tra file storage_state.json đã tồn tại chưa (tức là đã login)."""
    try:
        from notebooklm.paths import get_storage_path
        return get_storage_path().exists()
    except Exception:
        return False

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QListWidget, QListWidgetItem, QTextEdit, QLineEdit,
    QSplitter, QFileDialog, QMessageBox, QTabWidget, QWidget,
    QComboBox, QTreeWidget, QTreeWidgetItem,
)
from PySide6.QtCore import Qt, QThread, Signal, QTimer, QObject
from PySide6.QtGui import QFont


def _run_async(coro):
    """Chạy coroutine trong thread hiện tại."""
    return asyncio.run(coro)


def _patch_notebooklm_infographic():
    """Patch notebooklm library to fix infographic URL parsing.

    The library's _find_infographic_url uses a hard-coded path that may not match
    the current API response structure, causing wait_for_completion to time out
    even when the server marks the artifact as COMPLETED.

    Fixes:
    1. _is_media_ready → trusts COMPLETED status for infographic (no URL check).
    2. _find_infographic_url → adds a recursive fallback URL search.
    3. download_infographic patched to use the improved URL finder.
    """
    try:
        from notebooklm._artifacts import ArtifactsAPI, ArtifactTypeCode

        # ── Patch 1: trust COMPLETED for infographic ──────────────────────
        _orig_media_ready = ArtifactsAPI._is_media_ready

        def _patched_is_media_ready(self, art, artifact_type):
            if artifact_type == ArtifactTypeCode.INFOGRAPHIC.value:
                return True  # trust API status code; URL may parse later
            return _orig_media_ready(self, art, artifact_type)

        ArtifactsAPI._is_media_ready = _patched_is_media_ready

        # ── Patch 2: robust recursive URL search ──────────────────────────
        _orig_find_url = ArtifactsAPI._find_infographic_url

        def _deep_find_url(obj, depth=0):
            """Recursively search for an image-like http URL."""
            if depth > 14:
                return None
            if isinstance(obj, str) and obj.startswith(("http://", "https://")):
                low = obj.lower()
                if any(x in low for x in ('.png', '.jpg', '.jpeg', '.webp',
                                           'image/', 'ais.google', 'lh3.google',
                                           'storage.googleapis')):
                    return obj
            if isinstance(obj, list):
                for item in reversed(obj):
                    r = _deep_find_url(item, depth + 1)
                    if r:
                        return r
            return None

        def _patched_find_infographic_url(self, art):
            url = _orig_find_url(self, art)
            if url:
                return url
            return _deep_find_url(art)

        ArtifactsAPI._find_infographic_url = _patched_find_infographic_url

        # ── Patch 3: download_infographic uses patched _find_infographic_url ─
        import asyncio as _asyncio
        from notebooklm._artifacts import ArtifactStatus, ArtifactTypeCode as _ATC

        async def _patched_download_infographic(
                self, notebook_id, output_path, artifact_id=None):
            artifacts_data = await self._list_raw(notebook_id)
            info_candidates = [
                a for a in artifacts_data
                if isinstance(a, list) and len(a) > 4
                and a[2] == _ATC.INFOGRAPHIC
                and a[4] == ArtifactStatus.COMPLETED
            ]
            if artifact_id:
                info_art = next((i for i in info_candidates if i[0] == artifact_id), None)
            else:
                info_art = info_candidates[0] if info_candidates else None
            if not info_art:
                from notebooklm._artifacts import ArtifactNotReadyError
                raise ArtifactNotReadyError("infographic", artifact_id=artifact_id)
            url = self._find_infographic_url(info_art)
            if not url:
                from notebooklm._artifacts import ArtifactParseError
                raise ArtifactParseError("infographic",
                                         details="Could not find image URL")
            return await self._download_url(url, output_path)

        ArtifactsAPI.download_infographic = _patched_download_infographic

    except Exception:
        pass  # silent – never break the app


_patch_notebooklm_infographic()


# ── Workers ──────────────────────────────────────────────────────

class LoginWorker(QThread):
    done  = Signal()
    error = Signal(str)

    def __init__(self):
        super().__init__()
        self._event = threading.Event()

    def run(self):
        try:
            # Restore DefaultEventLoopPolicy for Playwright on Windows
            if sys.platform == "win32":
                asyncio.set_event_loop_policy(asyncio.DefaultEventLoopPolicy())

            from playwright.sync_api import sync_playwright
            from notebooklm.paths import get_storage_path, get_browser_profile_dir

            storage_path   = get_storage_path()
            browser_profile = get_browser_profile_dir()
            storage_path.parent.mkdir(parents=True, exist_ok=True)
            browser_profile.mkdir(parents=True, exist_ok=True)

            with sync_playwright() as p:
                context = p.chromium.launch_persistent_context(
                    user_data_dir=str(browser_profile),
                    headless=False,
                    args=[
                        "--disable-blink-features=AutomationControlled",
                        "--password-store=basic",
                    ],
                    ignore_default_args=["--enable-automation"],
                )
                page = context.pages[0] if context.pages else context.new_page()
                page.goto("https://notebooklm.google.com/")

                # Chờ user bấm "Save Login"
                self._event.wait()

                # Force .google.com cookies
                page.goto("https://accounts.google.com/", wait_until="load")
                page.goto("https://notebooklm.google.com/", wait_until="load")

                context.storage_state(path=str(storage_path))
                try:
                    storage_path.chmod(0o600)
                except Exception:
                    pass
                context.close()

            self.done.emit()
        except Exception as e:
            self.error.emit(str(e))
        finally:
            if sys.platform == "win32":
                asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

    def confirm(self):
        """Gửi tín hiệu lưu cookies sau khi đăng nhập xong."""
        self._event.set()


class ListNotebooksWorker(QThread):
    done  = Signal(list)
    error = Signal(str)

    def run(self):
        try:
            from notebooklm import NotebookLMClient
            async def _fetch():
                async with await NotebookLMClient.from_storage() as client:
                    return await client.notebooks.list()
            notebooks = _run_async(_fetch())
            self.done.emit(notebooks)
        except Exception as e:
            self.error.emit(str(e))


class CreateNotebookWorker(QThread):
    done  = Signal(object)
    error = Signal(str)

    def __init__(self, title: str):
        super().__init__()
        self.title = title

    def run(self):
        try:
            from notebooklm import NotebookLMClient
            async def _create():
                async with await NotebookLMClient.from_storage() as client:
                    return await client.notebooks.create(title=self.title)
            nb = _run_async(_create())
            self.done.emit(nb)
        except Exception as e:
            self.error.emit(str(e))


class DeleteNotebookWorker(QThread):
    done  = Signal()
    error = Signal(str)

    def __init__(self, notebook_id: str):
        super().__init__()
        self.notebook_id = notebook_id

    def run(self):
        try:
            from notebooklm import NotebookLMClient
            async def _delete():
                async with await NotebookLMClient.from_storage() as client:
                    await client.notebooks.delete(self.notebook_id)
            _run_async(_delete())
            self.done.emit()
        except Exception as e:
            self.error.emit(str(e))


class RenameNotebookWorker(QThread):
    done  = Signal(str)   # new title
    error = Signal(str)

    def __init__(self, notebook_id: str, new_title: str):
        super().__init__()
        self.notebook_id = notebook_id
        self.new_title   = new_title

    def run(self):
        try:
            from notebooklm import NotebookLMClient
            async def _rename():
                async with await NotebookLMClient.from_storage() as client:
                    nb = await client.notebooks.rename(self.notebook_id, self.new_title)
                    return getattr(nb, "title", self.new_title)
            title = _run_async(_rename())
            self.done.emit(title)
        except Exception as e:
            self.error.emit(str(e))


class ListSourcesWorker(QThread):
    done  = Signal(list)
    error = Signal(str)

    def __init__(self, notebook_id: str):
        super().__init__()
        self.notebook_id = notebook_id

    def run(self):
        try:
            from notebooklm import NotebookLMClient
            async def _fetch():
                async with await NotebookLMClient.from_storage() as client:
                    return await client.sources.list(self.notebook_id)
            sources = _run_async(_fetch())
            self.done.emit(sources)
        except Exception as e:
            self.error.emit(str(e))


class DeleteSourceWorker(QThread):
    done  = Signal()
    error = Signal(str)

    def __init__(self, notebook_id: str, source_id: str):
        super().__init__()
        self.notebook_id = notebook_id
        self.source_id   = source_id

    def run(self):
        try:
            from notebooklm import NotebookLMClient
            async def _delete():
                async with await NotebookLMClient.from_storage() as client:
                    await client.sources.delete(self.notebook_id, self.source_id)
            _run_async(_delete())
            self.done.emit()
        except Exception as e:
            self.error.emit(str(e))


class AddSourceWorker(QThread):
    done  = Signal()
    error = Signal(str)

    def __init__(self, notebook_id: str, file_path: str, use_vision: bool = False):
        super().__init__()
        self.notebook_id = notebook_id
        self.file_path   = file_path
        self.use_vision  = use_vision

    @staticmethod
    def _doc_to_docx(doc_path: str) -> str:
        """Convert .doc → .docx bằng Word COM, trả về path file tạm."""
        import tempfile, win32com.client as win32
        word = win32.Dispatch("Word.Application")
        word.Visible = False
        tmp = tempfile.mktemp(suffix=".docx")
        try:
            doc = word.Documents.Open(os.path.abspath(doc_path))
            doc.SaveAs(tmp, FileFormat=16)  # 16 = wdFormatXMLDocument (.docx)
            doc.Close(False)
        finally:
            word.Quit()
        return tmp

    # ── Vision helpers (Groq / Gemini / OpenRouter) ──────────────

    _VISION_PROMPT = """You are an expert engineering document analyst specializing in technical diagrams.
Analyze this page from an engineering document and provide a COMPLETE, STRUCTURED description.

IDENTIFY THE DIAGRAM TYPE FIRST:
- Function Block Diagram (FBD) / PLC Logic
- P&ID (Piping & Instrumentation Diagram)
- Electrical / Wiring Diagram
- Control Loop Diagram
- Process Flow Diagram (PFD)
- Schematic / Circuit Diagram
- Or plain text/table page

For EACH element found, extract:

=== INSTRUMENTS & TAGS ===
List every tag/instrument found. Format: TAG | TYPE | DESCRIPTION
Example: 10HFE30CP001 | Flow Transmitter | Measures hot primary air pressure

=== FUNCTION BLOCKS ===
List every function block/logic element. Format: BLOCK | FUNCTION | INPUTS | OUTPUT
Example: F(t) | Time-domain function | Signal from 10HFE30CP001 | Processed signal to Sigma block

=== SIGNAL FLOW & CONNECTIONS ===
Describe the data/signal path step by step, following arrows:
Step 1: [Source tag/block] → [Next block] — purpose/description
Step 2: [Block] → [Block] — ...
(Continue for all connections)

=== CONTROL LOGIC ===
- Control type (PID, cascade, feedforward, on-off, ...)
- Setpoints, process variables, manipulated variables
- Feedback loops: describe each loop
- Interlocks, permissives, or safety logic

=== PARAMETERS & SETPOINTS ===
List ALL numeric values: gains (K), time constants (T), setpoints, limits, delays
Format: PARAMETER | VALUE | UNIT | TAG/BLOCK

=== ALARMS & TRIPS ===
List any alarm/trip conditions visible

=== NOTES ===
Any additional labels, annotations, or notes on the page

IMPORTANT RULES:
- Use the EXACT tag names/numbers as shown (e.g. 10HFE30CP001X001, not "flow tag")
- If text is unclear, mark as [UNCLEAR: best guess]
- Do NOT skip any element, even if partially visible
- Write N/A for sections with no content"""

    @staticmethod
    def _call_openai_compat_vision(
        base_url: str, api_key: str, model: str,
        prompt: str, img_b64: str, timeout: int = 90,
    ) -> str:
        """G\u1ecdi b\u1ea5t k\u1ef3 endpoint OpenAI-compatible v\u1edbi Vision (base64 PNG)."""
        import requests
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        }
        payload = {
            "model": model,
            "messages": [{
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url",
                     "image_url": {"url": f"data:image/png;base64,{img_b64}"}},
                ],
            }],
            "temperature": 0.1,
            "max_tokens": 2048,
        }
        r = requests.post(
            f"{base_url.rstrip('/')}/chat/completions",
            headers=headers, json=payload, timeout=timeout,
        )
        r.raise_for_status()
        data = r.json()
        return (data["choices"][0]["message"]["content"] or "").strip()

    @staticmethod
    def _call_gemini_vision(
        api_key: str, model: str, prompt: str, img_b64: str, timeout: int = 90,
    ) -> str:
        """G\u1ecdi Gemini REST API v\u1edbi Vision (base64 PNG)."""
        import requests
        payload = {
            "contents": [{"parts": [
                {"text": prompt},
                {"inline_data": {"mime_type": "image/png", "data": img_b64}},
            ]}],
            "generationConfig": {"temperature": 0.1, "maxOutputTokens": 2048},
        }
        url = (
            f"https://generativelanguage.googleapis.com/v1beta/models/"
            f"{model}:generateContent?key={api_key}"
        )
        r = requests.post(url, json=payload, timeout=timeout)
        r.raise_for_status()
        data = r.json()
        return data["candidates"][0]["content"]["parts"][0]["text"].strip()

    def _extract_vision(self, pdf_path: str) -> str:
        """
        Phan tich tai lieu ky thuat bang Vision AI.
        - Vision AI xu ly TAT CA trang (khong loc theo word count)
        - fitz text chi la fallback neu Vision fail
        Thu tu fallback: Groq -> Gemini -> OpenRouter.
        """
        from core.llm_config import load_llm_config
        cfg = load_llm_config()

        groq_key       = (cfg.get("groq_api_key")      or "").strip()
        gemini_key     = (cfg.get("gemini_api_key")    or "").strip()
        openrouter_key = (cfg.get("openrouter_api_key") or "").strip()

        providers = []

        if groq_key:
            providers.append((
                "Groq (llama-4-scout)",
                lambda b64: self._call_openai_compat_vision(
                    base_url="https://api.groq.com/openai/v1",
                    api_key=groq_key,
                    model="meta-llama/llama-4-scout-17b-16e-instruct",
                    prompt=self._VISION_PROMPT,
                    img_b64=b64,
                ),
            ))

        if gemini_key:
            gemini_model = (cfg.get("gemini_model") or "gemini-2.0-flash").strip()
            providers.append((
                f"Gemini ({gemini_model})",
                lambda b64, m=gemini_model: self._call_gemini_vision(
                    api_key=gemini_key,
                    model=m,
                    prompt=self._VISION_PROMPT,
                    img_b64=b64,
                ),
            ))

        if openrouter_key:
            providers.append((
                "OpenRouter (llama-4-scout)",
                lambda b64: self._call_openai_compat_vision(
                    base_url="https://openrouter.ai/api/v1",
                    api_key=openrouter_key,
                    model="meta-llama/llama-4-scout:free",
                    prompt=self._VISION_PROMPT,
                    img_b64=b64,
                ),
            ))

        if not providers:
            raise ValueError(
                "No API key configured.\n"
                "Go to DB Search -> Config LLM to enter Groq / Gemini / OpenRouter key."
            )

        import fitz as _fitz, base64 as _b64

        # Render TAT CA trang thanh anh PNG, luu fitz text lam fallback
        doc = _fitz.open(pdf_path)
        pages_b64  = []
        pages_fitz = []
        for page in doc:
            pages_fitz.append(page.get_text("text").strip())
            pix = page.get_pixmap(dpi=200)
            if pix.colorspace and pix.colorspace.n != 3:
                pix = _fitz.Pixmap(_fitz.csRGB, pix)
            pages_b64.append(_b64.b64encode(pix.tobytes("png")).decode())
        doc.close()

        results = []
        used_provider = ""
        last_err = ""

        # Vision AI xu ly TAT CA trang — fitz chi la fallback neu Vision fail
        for i, (img_b64, fitz_text) in enumerate(zip(pages_b64, pages_fitz)):
            page_text = ""
            for label, try_fn in providers:
                try:
                    page_text = try_fn(img_b64)
                    if not used_provider:
                        used_provider = label
                    break
                except Exception as e:
                    last_err = f"{label}: {e}"
                    continue

            if page_text:
                results.append(f"--- Page {i+1} ---\n{page_text}")
            elif fitz_text:
                results.append(f"--- Page {i+1} [Text fallback] ---\n{fitz_text}")

        if not results and last_err:
            raise RuntimeError(
                f"Vision failed across all providers.\nLast error: {last_err}"
            )

        header = f"[Vision AI via {used_provider} -- {len(results)} pages]\n\n" if used_provider else ""
        return header + "\n\n".join(results)
    def run(self):
        tmp_file     = None
        vis_tmp_file = None
        try:
            from notebooklm import NotebookLMClient
            from pathlib import Path
            import tempfile

            upload_path = self.file_path
            if self.file_path.lower().endswith(".doc"):
                tmp_file = self._doc_to_docx(self.file_path)
                upload_path = tmp_file

            is_pdf = self.file_path.lower().endswith(".pdf")
            original_name = os.path.basename(self.file_path)

            # ── Vision AI (Groq / Gemini / OpenRouter auto-fallback) ──
            if self.use_vision and is_pdf:
                vision_str = self._extract_vision(self.file_path)
                if vision_str.strip():
                    vis_tmp_file = tempfile.mktemp(
                        prefix=f"[Vision] {original_name}_", suffix=".txt"
                    )
                    with open(vis_tmp_file, "w", encoding="utf-8") as f:
                        f.write(
                            f"Vision AI Analysis for: {original_name}\n\n"
                            f"{vision_str}"
                        )

            async def _add():
                async with await NotebookLMClient.from_storage() as client:
                    await client.sources.add_file(
                        self.notebook_id, Path(upload_path), wait=True
                    )
                    if vis_tmp_file:
                        await client.sources.add_file(
                            self.notebook_id, Path(vis_tmp_file), wait=True
                        )

            _run_async(_add())

            if not self.file_path.lower().endswith(".doc"):
                _upsert_source_map(
                    original_name,
                    self.file_path,
                    self.notebook_id,
                )
            self.done.emit()
        except Exception as e:
            self.error.emit(str(e))
        finally:
            for f in (tmp_file, vis_tmp_file):
                if f and os.path.exists(f):
                    try:
                        os.remove(f)
                    except Exception:
                        pass


class MergeNotebooksWorker(QThread):
    """Gộp nội dung nhiều notebook thành 1 notebook mới.

    Quy trình:
    1. Đọc nlm_source_map.json → tìm tất cả file thuộc các notebook đã check
    2. Extract text từ file local (PDF/DOCX/TXT) — không cần DB, không cần API
    3. Gộp thành file(s) .txt (~400k words/phần, dưới giới hạn NotebookLM)
    4. Tạo notebook mới và upload
    """
    progress         = Signal(str)      # status message
    notebook_created = Signal(str, str) # nb_id, nb_title
    done             = Signal(int, int) # n_files, n_parts
    error            = Signal(str)

    WORDS_PER_PART = 400_000

    def __init__(self, nb_map: dict, new_title: str):
        super().__init__()
        self.nb_map    = nb_map      # {nb_id: nb_title}
        self.new_title = new_title

    def _read_file(self, path: str) -> str:
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext in (".txt", ".md", ".csv"):
                with open(path, "r", encoding="utf-8", errors="ignore") as f:
                    return f.read()
            if ext == ".pdf":
                import pdfplumber
                with pdfplumber.open(path) as pdf:
                    return "\n".join(p.extract_text() or "" for p in pdf.pages)
            if ext in (".docx",):
                from docx import Document
                return "\n".join(p.text for p in Document(path).paragraphs)
        except Exception:
            pass
        return ""

    def run(self):
        import tempfile
        from pathlib import Path as _Path
        from notebooklm import NotebookLMClient

        try:
            # ── Bước 1: Tìm file thuộc các notebook đã check qua source map ──
            self.progress.emit("🔍 Đang tìm file từ source map…")
            src_map = _load_source_map()
            nb_ids  = set(self.nb_map.keys())

            # Lọc file có notebook_id khớp với danh sách đã check
            files_to_merge = []  # [(fname, path), ...]
            for fname, entry in src_map.items():
                if not isinstance(entry, dict):
                    continue
                nb_list = entry.get("notebooks", [])
                if any(nid in nb_ids for nid in nb_list):
                    path = entry.get("path", "")
                    if path and os.path.isfile(path):
                        files_to_merge.append((fname, path))

            if not files_to_merge:
                self.error.emit(
                    "Không tìm thấy file local nào thuộc các notebook đã check.\n"
                    "Hãy đảm bảo file vẫn còn trên đĩa và đã upload qua app này."
                )
                return

            total = len(files_to_merge)
            self.progress.emit(f"📄 Tìm thấy {total} file, đang đọc nội dung…")

            # ── Bước 2: Đọc nội dung từng file ───────────────────────────
            parts_data = []
            for i, (fname, path) in enumerate(files_to_merge, 1):
                if i % 50 == 0 or i == total:
                    self.progress.emit(f"📖 Đọc file {i}/{total}…")
                content = self._read_file(path)
                if content.strip():
                    parts_data.append((fname, content))

            if not parts_data:
                self.error.emit("Không đọc được nội dung từ bất kỳ file nào.")
                return

            self.progress.emit(f"✍️ Đang gộp {len(parts_data)} file thành text…")

            # ── Bước 3: Gộp và chia theo WORDS_PER_PART ──────────────────
            full_text = ""
            for fname, content in parts_data:
                full_text += f"\n\n{'='*60}\n📄 {fname}\n{'='*60}\n\n{content}"

            words   = full_text.split()
            n_parts = max(1, (len(words) + self.WORDS_PER_PART - 1) // self.WORDS_PER_PART)
            chunks  = [
                " ".join(words[i * self.WORDS_PER_PART:(i + 1) * self.WORDS_PER_PART])
                for i in range(n_parts)
            ]

            self.progress.emit(
                f"📤 Tạo notebook '{self.new_title}' và upload {n_parts} phần…"
            )

            # ── Bước 4: Tạo notebook và upload ───────────────────────────
            async def _create_and_upload():
                tmp_files = []
                try:
                    async with await NotebookLMClient.from_storage() as client:
                        nb    = await client.notebooks.create(title=self.new_title)
                        nb_id = getattr(nb, "id", None) or getattr(nb, "notebook_id", "")
                        for i, chunk in enumerate(chunks, 1):
                            tmp = tempfile.NamedTemporaryFile(
                                mode="w", suffix=".txt", encoding="utf-8",
                                delete=False, prefix="nlm_merge_"
                            )
                            tmp.write(chunk)
                            tmp.close()
                            tmp_files.append(tmp.name)
                            self.progress.emit(f"⬆ Upload phần {i}/{n_parts}…")
                            await client.sources.add_file(nb_id, _Path(tmp.name), wait=True)
                        return nb_id
                finally:
                    for f in tmp_files:
                        try:
                            os.remove(f)
                        except Exception:
                            pass

            nb_id = _run_async(_create_and_upload())
            self.notebook_created.emit(nb_id, self.new_title)
            self.done.emit(len(parts_data), n_parts)

        except Exception as e:
            self.error.emit(str(e))


class BatchAddSourceWorker(QThread):
    """Upload toàn bộ file trong folder lên NotebookLM.

    Cứ 100 file tạo một notebook mới tên '<folder> (Part N)'.
    Tự động làm mới session giữa các batch và retry khi gặp lỗi auth.
    """
    BATCH_SIZE   = 100
    CONCURRENCY  = 5   # số file upload song song trong 1 batch
    SUPPORTED    = {'.pdf', '.docx', '.doc', '.txt', '.xlsx', '.pptx', '.csv', '.md'}

    progress          = Signal(int, int, str, str)  # idx, total, filename, nb_title
    notebook_created  = Signal(str, str)             # nb_id, nb_title
    file_error        = Signal(str, str)             # filename, error_msg
    relogin_status    = Signal(str)                  # status message khi tự relogin
    done              = Signal(int, int, int)        # success, fail, skipped

    def __init__(self, folder_name: str, file_paths: list):
        super().__init__()
        self.folder_name = folder_name
        self.file_paths  = file_paths

    @staticmethod
    def _is_auth_error(err_msg: str) -> bool:
        """Kiểm tra xem lỗi có phải do session hết hạn / mất login không."""
        keywords = [
            "401", "403", "unauthorized", "authentication", "login",
            "session", "expired", "cookie", "credential", "auth",
            "not logged", "signed out", "access denied", "forbidden",
            "unauthenticated", "token", "permission",
        ]
        msg_lower = err_msg.lower()
        return any(kw in msg_lower for kw in keywords)

    @staticmethod
    def _auto_relogin() -> bool:
        """Làm mới session NotebookLM tự động bằng persistent browser context.

        Dùng user_data_dir đã lưu — Google cookies vẫn còn hiệu lực nên
        không cần tương tác người dùng. Trả True nếu thành công.
        """
        import time
        try:
            if sys.platform == "win32":
                asyncio.set_event_loop_policy(asyncio.DefaultEventLoopPolicy())

            from playwright.sync_api import sync_playwright
            from notebooklm.paths import get_storage_path, get_browser_profile_dir

            storage_path    = get_storage_path()
            browser_profile = get_browser_profile_dir()
            storage_path.parent.mkdir(parents=True, exist_ok=True)
            browser_profile.mkdir(parents=True, exist_ok=True)

            with sync_playwright() as p:
                context = p.chromium.launch_persistent_context(
                    user_data_dir=str(browser_profile),
                    headless=True,
                    args=[
                        "--disable-blink-features=AutomationControlled",
                        "--password-store=basic",
                        "--no-sandbox",
                        "--disable-dev-shm-usage",
                    ],
                    ignore_default_args=["--enable-automation"],
                )
                page = context.pages[0] if context.pages else context.new_page()
                try:
                    page.goto(
                        "https://notebooklm.google.com/",
                        wait_until="networkidle",
                        timeout=30_000,
                    )
                except Exception:
                    page.goto("https://notebooklm.google.com/")
                time.sleep(2)

                # Nếu trang yêu cầu đăng nhập lại (headless không thể làm), dùng visible browser
                if "accounts.google.com" in page.url or "signin" in page.url.lower():
                    context.close()
                    # Fallback: visible browser, tự động vì profile đã có session
                    context2 = p.chromium.launch_persistent_context(
                        user_data_dir=str(browser_profile),
                        headless=False,
                        args=[
                            "--disable-blink-features=AutomationControlled",
                            "--password-store=basic",
                        ],
                        ignore_default_args=["--enable-automation"],
                    )
                    page2 = context2.pages[0] if context2.pages else context2.new_page()
                    page2.goto("https://notebooklm.google.com/", wait_until="load", timeout=60_000)
                    time.sleep(3)
                    context2.storage_state(path=str(storage_path))
                    try:
                        storage_path.chmod(0o600)
                    except Exception:
                        pass
                    context2.close()
                else:
                    context.storage_state(path=str(storage_path))
                    try:
                        storage_path.chmod(0o600)
                    except Exception:
                        pass
                    context.close()

            return True
        except Exception:
            return False
        finally:
            if sys.platform == "win32":
                asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

    def run(self):
        from notebooklm import NotebookLMClient
        from pathlib import Path as _Path

        # Lọc file đã upload — đồng thời đếm notebook hiện có để đánh số đúng
        src_map          = _load_source_map()
        new_files        = []
        skipped          = 0
        existing_nb_ids  = set()  # notebook đã chứa file của folder này

        for fp in self.file_paths:
            fname = os.path.basename(fp)
            entry = src_map.get(fname, {})
            if isinstance(entry, dict) and entry.get("notebooks"):
                skipped += 1
                for nb_id in entry.get("notebooks", []):
                    existing_nb_ids.add(nb_id)
            else:
                new_files.append(fp)

        # Số notebook đã tạo trước → offset cho phần đánh số
        part_offset = len(existing_nb_ids)

        success = fail = 0
        total   = len(new_files)

        if not new_files:
            self.done.emit(0, 0, skipped)
            return

        batches = [new_files[i:i + self.BATCH_SIZE]
                   for i in range(0, total, self.BATCH_SIZE)]

        def _nb_title_for(batch_idx: int) -> str:
            part_num = part_offset + batch_idx
            if part_num == 0:
                return self.folder_name
            return f"{self.folder_name} (Part {part_num + 1})"

        def _create_notebook(title: str) -> str:
            async def _c(t=title):
                async with await NotebookLMClient.from_storage() as client:
                    nb = await client.notebooks.create(title=t)
                    return getattr(nb, "id", None) or getattr(nb, "notebook_id", "")
            return _run_async(_c())

        def _upload_file(path: str, nid: str):
            async def _a(p=path, n=nid):
                async with await NotebookLMClient.from_storage() as client:
                    await client.sources.add_file(n, _Path(p), wait=True)
            _run_async(_a())

        for batch_idx, batch in enumerate(batches):
            # ── Làm mới session trước khi bắt đầu batch (trừ batch đầu tiên) ──
            if batch_idx > 0:
                self.relogin_status.emit("🔄 Đang làm mới session Google...")
                ok = self._auto_relogin()
                self.relogin_status.emit(
                    "✅ Session đã làm mới, tiếp tục upload..." if ok
                    else "⚠ Không thể tự làm mới session, thử tiếp..."
                )

            nb_title = _nb_title_for(batch_idx)

            # ── Tạo notebook (retry 1 lần nếu lỗi auth) ──
            nb_id = None
            try:
                nb_id = _create_notebook(nb_title)
                self.notebook_created.emit(nb_id, nb_title)
            except Exception as e:
                if self._is_auth_error(str(e)):
                    self.relogin_status.emit("🔑 Session hết hạn khi tạo notebook, đang đăng nhập lại...")
                    if self._auto_relogin():
                        try:
                            nb_id = _create_notebook(nb_title)
                            self.notebook_created.emit(nb_id, nb_title)
                        except Exception as e2:
                            self.file_error.emit(f"[Tạo notebook '{nb_title}']", str(e2))
                            fail += len(batch)
                            continue
                    else:
                        self.file_error.emit(f"[Tạo notebook '{nb_title}']", str(e))
                        fail += len(batch)
                        continue
                else:
                    self.file_error.emit(f"[Tạo notebook '{nb_title}']", str(e))
                    fail += len(batch)
                    continue

            # ── Pre-process: convert .doc → .docx (sync, trước khi upload) ──
            upload_items = []   # list of (fp, fname, upload_path, tmp_file)
            for fp in batch:
                fname = os.path.basename(fp)
                tmp_file, upload_path = None, fp
                if fp.lower().endswith(".doc"):
                    try:
                        tmp_file    = AddSourceWorker._doc_to_docx(fp)
                        upload_path = tmp_file
                    except Exception:
                        pass
                upload_items.append((fp, fname, upload_path, tmp_file))

            # ── Upload tuần tự trên 1 client (an toàn với mọi loại library) ──
            _b_idx   = batch_idx   # capture cho closure
            _nb_ttl  = nb_title
            _total   = total

            async def _upload_sequential(items, nid):
                from notebooklm import NotebookLMClient as _NLC
                results = []
                async with await _NLC.from_storage() as client:
                    for i, (fp, fname, upload_path, _) in enumerate(items):
                        try:
                            await client.sources.add_file(nid, _Path(upload_path), wait=True)
                            results.append((fp, fname, None))
                        except Exception as e:
                            results.append((fp, fname, e))
                        # Emit progress ngay sau mỗi file (không chờ cả batch xong)
                        g_idx = (_b_idx * self.BATCH_SIZE) + i + 1
                        self.progress.emit(g_idx, _total, fname, _nb_ttl)
                return results

            raw_results = _run_async(_upload_sequential(upload_items, nb_id))

            # raw_results: [(fp, fname, error_or_None), ...]
            tmp_map = {fp: tmp_file for fp, fname, upload_path, tmp_file in upload_items}
            up_map  = {fp: upload_path for fp, fname, upload_path, tmp_file in upload_items}

            for res_fp, res_fname, err in raw_results:

                if err is not None:
                    err_msg = str(err)
                    if self._is_auth_error(err_msg):
                        self.relogin_status.emit(f"🔑 Auth expired khi upload '{res_fname}', đang refresh...")
                        if self._auto_relogin():
                            try:
                                _upload_file(up_map[res_fp], nb_id)
                                _upsert_source_map(res_fname, res_fp, nb_id)
                                success += 1
                            except Exception as e2:
                                self.file_error.emit(res_fname, str(e2))
                                fail += 1
                        else:
                            self.file_error.emit(res_fname, err_msg)
                            fail += 1
                    else:
                        self.file_error.emit(res_fname, err_msg)
                        fail += 1
                else:
                    _upsert_source_map(res_fname, res_fp, nb_id)
                    success += 1

                tmp_file = tmp_map.get(res_fp)
                if tmp_file and os.path.exists(tmp_file):
                    try:
                        os.remove(tmp_file)
                    except Exception:
                        pass

        self.done.emit(success, fail, skipped)


_MINDMAP_PROMPT = """You are a technical document analyst. Read the ENTIRE document thoroughly and produce a COMPREHENSIVE mind map in strict JSON format.

RULES:
1. Return ONLY a single JSON object — no explanation, no markdown, no code fences.
2. Every node has: "name" (string), "children" (array, can be empty).
3. Leaf nodes (no children) MUST also have:
   - "description": exact quote or close paraphrase from the document (max 2 sentences)
   - "page": page number as integer (0 if unknown)
4. Group nodes (have children) should NOT have "description".
5. Max depth: 4 levels. Root = document title.
6. Keep "name" short (≤ 10 words). Put details in "description".

COMPLETENESS RULES (CRITICAL):
- Extract EVERY distinct topic, system, subsystem, and component mentioned.
- Extract ALL numerical values: temperatures, pressures, flows, speeds, voltages, timings, tolerances, clearances.
- Extract ALL alarm/trip setpoints, normal/allowable/emergency limits.
- Extract ALL procedures and their steps (each step = one leaf node).
- Extract ALL warnings, cautions, and notes.
- Extract ALL maintenance intervals, inspection criteria, acceptance criteria.
- Do NOT skip any section. Do NOT merge different items into one node.
- If a section has 10 items, create 10 leaf nodes — not a summary.

Example format:
{
  "name": "Document Title",
  "children": [
    {
      "name": "System A",
      "children": [
        {
          "name": "Operating Temp: 27-35°C",
          "description": "Feed oil temperature shall be maintained between 27°C and 35°C during normal start-up.",
          "page": 12,
          "children": []
        }
      ]
    }
  ]
}

Now produce the COMPLETE mind map JSON for this document:"""

_MINDMAP_LANG_SUFFIX = {
    "vi": '\n7. Write ALL "name" and "description" text in Vietnamese.',
    "ko": '\n7. Write ALL "name" and "description" text in Korean.',
    "ja": '\n7. Write ALL "name" and "description" text in Japanese.',
}

def _build_mindmap_prompt(lang: str = "en") -> str:
    return _MINDMAP_PROMPT + _MINDMAP_LANG_SUFFIX.get(lang, "")

_SUMMARIZE_PROMPT = (
    "Please provide a comprehensive summary of this document following its original structure and headings. "
    "Include: main topics, key findings or procedures, important technical data, "
    "warnings or critical notices, and key takeaways. "
    "Use plain text only — no markdown, no ### or ** symbols, no bullet asterisks. "
    "Use the document's own section titles as headings."
)

_TRANSLATE_VI_PROMPT = (
    "Now provide a comprehensive summary of the same document in Vietnamese. "
    "Follow the same structure and section headings as the document. "
    "Include all main topics, key findings, procedures, technical data, warnings, and key takeaways — same level of detail as the English summary. "
    "Write in proper Vietnamese with full diacritical marks and tone marks (ă, â, đ, ê, ô, ơ, ư and all tone accents). "
    "Do NOT use ASCII transliteration (write 'không' not 'khong', 'được' not 'duoc', etc.). "
    "Use plain text only — no markdown, no ### or ** symbols, no bullet asterisks."
)


class TempChatWorker(QThread):
    """Tạo notebook tạm, upload file → emit (notebook_id, title) để mở chat."""
    done  = Signal(str, str)  # notebook_id, notebook_title
    error = Signal(str)

    def __init__(self, file_path: str):
        super().__init__()
        self.file_path = file_path

    def run(self):
        try:
            from notebooklm import NotebookLMClient
            fname = os.path.basename(self.file_path)
            title = f"[Chat] {fname}"
            async def _work():
                async with await NotebookLMClient.from_storage() as client:
                    nb = await client.notebooks.create(title=title)
                    nb_id = getattr(nb, "id", None) or getattr(nb, "notebook_id", "")
                    await client.sources.add_file(nb_id, Path(self.file_path), wait=True)
                    return nb_id, title
            nb_id, nb_title = _run_async(_work())
            self.done.emit(nb_id, nb_title)
        except Exception as e:
            self.error.emit(str(e))


class NLMAutoSummarizeWorker(QThread):
    """Auto-pick first notebook (or create one) → upload → summarize → delete source."""
    done  = Signal(str)
    error = Signal(str)

    def __init__(self, file_path: str):
        super().__init__()
        self.file_path = file_path

    def run(self):
        try:
            from notebooklm import NotebookLMClient
            async def _work():
                async with await NotebookLMClient.from_storage() as client:
                    notebooks = await client.notebooks.list()
                    if notebooks:
                        nb = notebooks[0]
                        nb_id = getattr(nb, "id", None) or getattr(nb, "notebook_id", "")
                    else:
                        nb = await client.notebooks.create(title="Auto Summary")
                        nb_id = getattr(nb, "id", None) or getattr(nb, "notebook_id", "")
                    source = await client.sources.add_file(nb_id, Path(self.file_path), wait=True)
                    sid = getattr(source, "source_id", None) or getattr(source, "id", None)
                    result_orig = await client.chat.ask(nb_id, _SUMMARIZE_PROMPT, source_ids=[sid])
                    summary_orig = (getattr(result_orig, "answer", None) or str(result_orig)).strip()
                    result_vi = await client.chat.ask(nb_id, _TRANSLATE_VI_PROMPT, source_ids=[sid])
                    summary_vi = (getattr(result_vi, "answer", None) or str(result_vi)).strip()
                    await client.sources.delete(nb_id, sid)
                    separator = "\n\n" + "─" * 48 + "\n🇻🇳  BẢN TIẾNG VIỆT\n" + "─" * 48 + "\n\n"
                    return summary_orig + separator + summary_vi
            text = _run_async(_work())
            self.done.emit(text)
        except Exception as e:
            self.error.emit(str(e))


_MINDMAP_TEMPLATE = r"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>__TITLE__</title>
<style>
  *{box-sizing:border-box;}
  html,body{margin:0;padding:0;background:#080c14;width:100%;height:100%;overflow:hidden;font-family:sans-serif;}
  #app{width:100%;height:100%;}
  svg{width:100%;height:100%;display:block;cursor:grab;}
  svg.dragging{cursor:grabbing;}
  .nd{cursor:pointer;}
  .nd:hover rect,.nd:hover ellipse{filter:brightness(1.25);}
  .lk{fill:none;}
  text{pointer-events:none;dominant-baseline:middle;}
  #toolbar{position:fixed;top:8px;left:50%;transform:translateX(-50%);display:flex;gap:5px;z-index:20;
    background:rgba(8,12,20,0.92);padding:6px 10px;border-radius:10px;
    border:1px solid rgba(255,255,255,0.1);backdrop-filter:blur(8px);}
  #toolbar button{background:rgba(255,255,255,0.07);color:#c8d4f0;border:1px solid rgba(255,255,255,0.13);
    border-radius:6px;padding:4px 10px;font-size:12px;cursor:pointer;transition:background .15s;}
  #toolbar button:hover{background:rgba(255,255,255,0.18);}
  #search{background:rgba(255,255,255,0.07);color:#c8d4f0;border:1px solid rgba(255,255,255,0.18);
    border-radius:6px;padding:4px 8px;font-size:12px;width:140px;outline:none;}
  #search::placeholder{color:rgba(200,212,240,0.35);}
  #detail{position:fixed;max-width:260px;min-width:180px;
    background:rgba(8,12,20,0.97);border:1px solid rgba(255,255,255,0.14);border-radius:12px;
    padding:12px 14px;color:#c8d4f0;font-size:12px;display:none;z-index:30;line-height:1.6;
    box-shadow:0 4px 24px rgba(0,0,0,0.6);}
  #detail .dtitle{font-weight:bold;font-size:13px;margin-bottom:6px;color:#60a5fa;word-break:break-word;}
  #detail .dbread{font-size:10px;color:rgba(200,212,240,0.4);margin-bottom:6px;}
  #detail .dchildren{margin-top:6px;font-size:11px;color:#a0b0d0;}
  #detail .dchildren div{padding:2px 0;border-top:1px solid rgba(255,255,255,0.06);}
  #detail .dquote{margin-top:6px;font-size:11px;color:#c8d4f0;font-style:italic;border-left:2px solid #3b82f6;padding-left:8px;}
  #detail .dpage{margin-top:4px;font-size:10px;color:rgba(200,212,240,0.4);}
  #detail .dclose{float:right;cursor:pointer;opacity:0.4;font-size:15px;line-height:1;margin-left:6px;color:#c8d4f0;}
  #detail .dclose:hover{opacity:1;}
</style></head><body>
<div id="toolbar">
  <input id="search" type="text" placeholder="🔍 Search…" oninput="doSearch(this.value)">
  <button onclick="fitScreen()">⊡ Fit</button>
  <button onclick="expandAll()">+ All</button>
  <button onclick="collapseDeep()">− Deep</button>
  <button onclick="collapseAll()">− All</button>

</div>
<div id="detail">
  <span class="dclose" onclick="closeDetail()">✕</span>
  <div class="dbread" id="det-bread"></div>
  <div class="dtitle" id="det-title"></div>
  <div class="dchildren" id="det-children"></div>
</div>
<div id="app"><svg id="svg"><g id="canvas"></g></svg></div>
<script>
const RAW=__TREE_DATA__;
const TITLE="__TITLE__";

// ── Topic palettes (one per level-1 branch) ─────────────
const PAL=[
  {node:'#1a3560',text:'#89b4fa',edge:'#4a80c8',accent:'#89b4fa'},
  {node:'#1a3d22',text:'#a6e3a1',edge:'#4caf70',accent:'#a6e3a1'},
  {node:'#3d2800',text:'#fab387',edge:'#d07030',accent:'#fab387'},
  {node:'#301a50',text:'#cba6f7',edge:'#8060c0',accent:'#cba6f7'},
  {node:'#3d1020',text:'#f38ba8',edge:'#c04060',accent:'#f38ba8'},
  {node:'#0d3530',text:'#94e2d5',edge:'#30a898',accent:'#94e2d5'},
  {node:'#3d3000',text:'#f9e2af',edge:'#c09030',accent:'#f9e2af'},
];

// ── Stats pattern: numbers with technical units ──────────
const STATS_RE=/\b\d[\d,.]* *(°[CF]|bar|rpm|%|kg|kw|mw|kv|hz|m\/s|cycles?|h\b|hr\b|min\b|ms\b|psi|mpa|kpa|kj|mj|kwh|mwh|°)\b/i;
function isStats(s){return STATS_RE.test(s||'');}

// ── Tree init ────────────────────────────────────────────
let uid=0;
function initTree(n,parent,level,topic){
  n._id=uid++;n._parent=parent;n._level=level;
  n._collapsed=(level>=3);
  n._topic=topic;
  n._detail=n.description||n.detail||n.content||'';
  n._page=n.page||0;
  n.children=(n.children||[]);
  if(level===0) n.children.forEach((c,i)=>initTree(c,n,1,i%PAL.length));
  else          n.children.forEach(c=>initTree(c,n,level+1,topic));
}
initTree(RAW,null,0,0);

// ── Leaf count (for arc allocation) ─────────────────────
function leaves(n){
  if(n._collapsed||!n.children.length) return 1;
  return n.children.reduce((s,c)=>s+leaves(c),0);
}

// ── Horizontal tree layout ───────────────────────────────
const NH=30,VGAP=8,HGAP=44;
let _bx0=0,_bx1=0,_by0=0,_by1=0;

function subtreeH(n){
  if(n._collapsed||!n.children.length) return NH+VGAP;
  return n.children.reduce((s,c)=>s+subtreeH(c),0);
}

function layout(n,x,y0,lv){
  n._lv=lv; n._x=x;
  if(n._collapsed||!n.children.length){n._y=y0;return NH+VGAP;}
  const tot=n.children.reduce((s,c)=>s+subtreeH(c),0);
  let cy=y0;
  n.children.forEach(c=>{layout(c,x+nw(n.name,lv)+HGAP,cy,lv+1);cy+=subtreeH(c);});
  n._y=y0+(tot-NH)/2;
  return tot;
}

function doLayout(){
  layout(RAW,24,24,0);
  const vis=allVis(RAW);
  _bx0=0; _bx1=Math.max(...vis.map(n=>n._x+nw(n.name,n._lv)))+24;
  _by0=0; _by1=Math.max(...vis.map(n=>n._y+NH))+24;
}

function allVis(n){
  let r=[n];
  if(!n._collapsed) n.children.forEach(c=>r=r.concat(allVis(c)));
  return r;
}

// ── Helpers ──────────────────────────────────────────────
function esc(s){return(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');}
function nw(name,lv){return Math.min(200,Math.max(64,(name||'').length*7+(lv===0?28:18)));}
function nh(lv){return lv===0?38:lv===1?32:28;}

let _q='';

// ── Render ───────────────────────────────────────────────
function render(){
  doLayout();
  const vis=allVis(RAW);
  let s='';

  // ── Edges ──
  vis.forEach(n=>{
    if(!n._parent) return;
    const p=n._parent;
    const ECOLS=['#3b82f6','#10b981','#f59e0b','#a855f7','#f43f5e','#14b8a6'];
    const ec=ECOLS[Math.min(n._lv-1,ECOLS.length-1)];
    const pw=nw(p.name,p._lv);
    const x1=p._x+pw, y1=p._y+NH/2;
    const x2=n._x,    y2=n._y+NH/2;
    const mx=(x1+x2)/2;
    const sw=n._lv<=1?2.5:1.5, op=n._lv<=1?0.7:0.45;
    s+=`<path class="lk" stroke="${ec}" stroke-width="${sw}" opacity="${op}"
      d="M${x1},${y1} C${mx},${y1} ${mx},${y2} ${x2},${y2}"/>`;
  });

  // ── Nodes ──
  vis.forEach(n=>{
    const w=nw(n.name,n._lv), h=nh(n._lv);
    const x=n._x, y=n._y;
    const stats=isStats(n.name);
    const matched=_q&&(n.name||'').toLowerCase().includes(_q);
    const hasKids=n.children.length>0;

    // Fill / stroke — color by level
    const LCOLS=[
      {fill:'#0f1b2d',text:'#60a5fa',stroke:'#3b82f6'},  // lv0 — sapphire
      {fill:'#0d2b1e',text:'#34d399',stroke:'#10b981'},  // lv1 — emerald
      {fill:'#2d1b00',text:'#fbbf24',stroke:'#f59e0b'},  // lv2 — amber
      {fill:'#1e0d35',text:'#c084fc',stroke:'#a855f7'},  // lv3 — violet
      {fill:'#2d0a14',text:'#fb7185',stroke:'#f43f5e'},  // lv4 — rose
      {fill:'#012a2a',text:'#2dd4bf',stroke:'#14b8a6'},  // lv5 — cyan
    ];
    const lc=LCOLS[Math.min(n._lv,LCOLS.length-1)];
    let fill=lc.fill,textC=lc.text,strokeC=lc.stroke;
    let strokeW=n._lv===0?2.5:n._lv===1?2:1.2;
    let rx=n._lv===0?16:12;
    if(stats){fill='#422006';textC='#fde68a';strokeC='#d97706';strokeW=2;rx=8;}
    if(matched){fill='#2a1e00';strokeC='#fbbf24';strokeW=2.5;textC='#fde68a';}

    const fs=n._lv===0?14:n._lv===1?12:11;
    const fw=n._lv<=1?'bold':'normal';
    const label=esc(n.name||'')+(n._collapsed&&hasKids?' ▸':'');

    s+=`<g class="nd" data-name="${esc(n.name||'')}" onclick="nodeClick(${n._id},event)">`;
    s+=`<rect x="${x}" y="${y}" width="${w}" height="${h}" rx="${rx}" fill="${fill}" stroke="${strokeC}" stroke-width="${strokeW}"/>`;
    // Stats badge
    if(stats) s+=`<circle cx="${x+w-7}" cy="${y+7}" r="4" fill="#c09030" opacity="0.9"/>`;
    // Detail dot
    if(n._detail) s+=`<circle cx="${x+7}" cy="${y+7}" r="3" fill="${lc.text}" opacity="0.7"/>`;
    s+=`<title>${esc(n.name||'')}${n._detail?'\n\n'+n._detail.slice(0,240):''}</title>`;
    s+=`<text x="${x+w/2}" y="${y+h/2}" text-anchor="middle" fill="${textC}" font-size="${fs}" font-weight="${fw}">${label}</text>`;
    s+=`</g>`;
  });

  document.getElementById('canvas').innerHTML=s;
}

// ── Node click ───────────────────────────────────────────
function nodeClick(id,evt){
  evt.stopPropagation();
  const vis=allVis(RAW);
  const n=vis.find(x=>x._id===id);
  if(!n) return;
  showDetail(n,evt);
  if(n.children.length){n._collapsed=!n._collapsed;render();}
}

function nodeRightClick(name,evt){
  evt.preventDefault();
  evt.stopPropagation();
  document.getElementById('ctx-menu')?.remove();
  const m=document.createElement('div');
  m.id='ctx-menu';
  m.style.cssText=`position:fixed;left:${evt.clientX}px;top:${evt.clientY}px;
    background:#313244;border:1px solid #45475a;border-radius:6px;
    padding:4px 0;z-index:999;box-shadow:0 4px 12px #0006;min-width:160px;`;
  const item=document.createElement('div');
  item.textContent='📓 Send to NbLM';
  item.style.cssText='padding:8px 14px;cursor:pointer;font-size:13px;color:#cdd6f4;';
  item.onmouseenter=()=>item.style.background='#45475a';
  item.onmouseleave=()=>item.style.background='';
  item.onclick=()=>{m.remove();if(name)console.log('mm:ask:'+name);};
  m.appendChild(item);
  document.body.appendChild(m);
  const close=()=>m.remove();
  setTimeout(()=>document.addEventListener('click',close,{once:true}),0);
}

function breadcrumb(n){
  const parts=[];
  let cur=n._parent;
  while(cur){parts.unshift(cur.name||'');cur=cur._parent;}
  return parts.join(' › ');
}

function showDetail(n,evt){
  const dlg=document.getElementById('detail');
  document.getElementById('det-title').textContent=n.name||'';
  const bread=breadcrumb(n);
  const breadEl=document.getElementById('det-bread');
  breadEl.textContent=bread; breadEl.style.display=bread?'block':'none';
  const chEl=document.getElementById('det-children');
  if(n.children.length){
    chEl.innerHTML=n.children.map(c=>`<div>▸ ${esc(c.name||'')}</div>`).join('');
    chEl.style.display='block';
  } else {
    chEl.innerHTML=''; chEl.style.display='none';
  }
  // Quote + page
  let quoteEl=document.getElementById('det-quote');
  if(!quoteEl){quoteEl=document.createElement('div');quoteEl.id='det-quote';dlg.appendChild(quoteEl);}
  let pageEl=document.getElementById('det-page');
  if(!pageEl){pageEl=document.createElement('div');pageEl.id='det-page';dlg.appendChild(pageEl);}
  if(n._detail){
    quoteEl.className='dquote'; quoteEl.textContent='"'+n._detail+'"'; quoteEl.style.display='block';
  } else { quoteEl.style.display='none'; }
  if(n._page){
    pageEl.className='dpage';
    pageEl.innerHTML='<a href="#" onclick="goToPage('+n._page+');return false;" style="color:#60a5fa;text-decoration:underline;cursor:pointer;">📄 Page '+n._page+'</a>';
    pageEl.style.display='block';
  } else { pageEl.style.display='none'; }
  // position near click, avoid overflow
  dlg.style.display='block';
  const sw=window.innerWidth, sh=window.innerHeight;
  const dw=dlg.offsetWidth||260, dh=dlg.offsetHeight||120;
  let px=evt.clientX+14, py=evt.clientY-20;
  if(px+dw>sw-8) px=evt.clientX-dw-14;
  if(py+dh>sh-8) py=sh-dh-8;
  if(py<8) py=8;
  dlg.style.left=px+'px'; dlg.style.top=py+'px';
}
function closeDetail(){document.getElementById('detail').style.display='none';}
function goToPage(pg){console.log('mm:page:'+pg);}

// ── Search ───────────────────────────────────────────────
function doSearch(val){
  _q=val.trim().toLowerCase();
  if(_q) expandForSearch(RAW,_q);
  render();
}
function expandForSearch(n,q){
  const mine=(n.name||'').toLowerCase().includes(q);
  const childHit=n.children.some(c=>expandForSearch(c,q));
  if(childHit) n._collapsed=false;
  return mine||childHit;
}

// ── Expand / Collapse ────────────────────────────────────
function setAll(n,v){n._collapsed=v;n.children.forEach(c=>setAll(c,v));}
function expandAll(){setAll(RAW,false);render();fitScreen();}
function collapseAll(){RAW.children.forEach(c=>setAll(c,true));render();fitScreen();}
function collapseDeep(){collapseLevel(RAW,2);render();fitScreen();}
function collapseLevel(n,max){n._collapsed=(n._lv>=max);n.children.forEach(c=>collapseLevel(c,max));}

// ── Zoom & Pan ───────────────────────────────────────────
const svg=document.getElementById('svg');
const canvas=document.getElementById('canvas');
let vx=0,vy=0,vs=1;
function applyT(){canvas.setAttribute('transform',`translate(${vx},${vy}) scale(${vs})`);}
function fitScreen(){
  const sw=svg.clientWidth,sh=svg.clientHeight;
  const cw=_bx1-_bx0,ch=_by1-_by0;
  if(cw<=0||ch<=0) return;
  vs=Math.min(sw/cw,sh/ch)*0.88;
  vx=sw/2-(_bx0+cw/2)*vs; vy=sh/2-(_by0+ch/2)*vs;
  applyT();
}
svg.addEventListener('wheel',e=>{
  e.preventDefault();
  const r=svg.getBoundingClientRect();
  const mx=e.clientX-r.left,my=e.clientY-r.top;
  const f=e.deltaY<0?1.12:0.89;
  const ns=Math.min(6,Math.max(0.08,vs*f));
  vx=mx-(mx-vx)*(ns/vs); vy=my-(my-vy)*(ns/vs); vs=ns; applyT();
},{passive:false});
let drag=false,ddx=0,ddy=0;
svg.addEventListener('mousedown',e=>{
  if(e.target.closest('.nd')) return;
  drag=true; ddx=e.clientX-vx; ddy=e.clientY-vy; svg.classList.add('dragging');
});
window.addEventListener('mousemove',e=>{if(!drag)return;vx=e.clientX-ddx;vy=e.clientY-ddy;applyT();});
window.addEventListener('mouseup',()=>{drag=false;svg.classList.remove('dragging');});
svg.addEventListener('click',()=>closeDetail());



// ── Keyboard shortcuts ───────────────────────────────────
document.addEventListener('keydown',e=>{
  if(e.key==='Escape'){closeDetail();document.getElementById('search').value='';doSearch('');}
  if(e.key==='f'&&!e.ctrlKey&&!e.metaKey) fitScreen();
});

// ── Context menu (right-click node → Send to NbLM) ───────
document.addEventListener('contextmenu', e=>{
  e.preventDefault();
  document.getElementById('ctx-menu')?.remove();
  const g = e.target.closest('g.nd');
  if(!g) return;
  const name = g.getAttribute('data-name');
  if(!name) return;
  const m = document.createElement('div');
  m.id = 'ctx-menu';
  m.style.cssText = `position:fixed;left:${e.clientX}px;top:${e.clientY}px;
    background:#313244;border:1px solid #45475a;border-radius:6px;
    padding:4px 0;z-index:9999;box-shadow:0 4px 12px #0008;min-width:170px;`;
  const item = document.createElement('div');
  item.textContent = '📓 Send to NbLM';
  item.style.cssText = 'padding:8px 14px;cursor:pointer;font-size:13px;color:#cdd6f4;white-space:nowrap;';
  item.onmouseenter = ()=>item.style.background='#45475a';
  item.onmouseleave = ()=>item.style.background='';
  item.onclick = ()=>{ m.remove(); console.log('mm:ask:'+name); };
  m.appendChild(item);
  document.body.appendChild(m);
  setTimeout(()=>document.addEventListener('click', ()=>m.remove(), {once:true}), 0);
});

// ── Init ─────────────────────────────────────────────────
render();
setTimeout(fitScreen,80);
</script></body></html>"""


def _mindmap_to_html(node: dict, title: str = "") -> str:
    """Convert mind map tree dict → self-contained SVG mind map HTML."""
    import json
    return (_MINDMAP_TEMPLATE
            .replace("__TITLE__", title)
            .replace("__TREE_DATA__", json.dumps(node, ensure_ascii=False)))


class MindMapWorker(QThread):
    """Upload file → generate_mind_map() → render HTML → delete source."""
    done  = Signal(str)   # HTML content
    error = Signal(str)

    def __init__(self, file_path: str, lang: str = "en"):
        super().__init__()
        self.file_path = file_path
        self.lang = lang

    def run(self):
        try:
            from notebooklm import NotebookLMClient
            import os
            async def _work():
                async with await NotebookLMClient.from_storage() as client:
                    # Use a dedicated temp notebook so only this file is in scope
                    title = os.path.basename(self.file_path)
                    nb = await client.notebooks.create(title=f"[MindMap] {title}")
                    nb_id = getattr(nb, "id", None) or getattr(nb, "notebook_id", "")
                    try:
                        source = await client.sources.add_file(nb_id, Path(self.file_path), wait=True)
                        sid = getattr(source, "source_id", None) or getattr(source, "id", None)
                        result = await client.chat.ask(nb_id, _build_mindmap_prompt(self.lang), source_ids=[sid] if sid else None)
                        raw_text = (getattr(result, "answer", None) or str(result)).strip()
                        # Parse JSON — strip any accidental markdown fences
                        import json, re
                        m = re.search(r'\{.*\}', raw_text, re.DOTALL)
                        tree = json.loads(m.group(0)) if m else {"name": title, "children": []}
                        return _mindmap_to_html(tree, title)
                    finally:
                        # Always clean up the temp notebook
                        try:
                            await client.notebooks.delete(nb_id)
                        except Exception:
                            pass
            html = _run_async(_work())
            self.done.emit(html)
        except Exception as e:
            self.error.emit(str(e))


class MindMapWorkerOnline(QThread):
    """Tạo mind map từ source đã có sẵn trên NbLM (không cần file local)."""
    done  = Signal(str)   # HTML content
    error = Signal(str)

    def __init__(self, notebook_id: str, title: str, source_id: str = "", lang: str = "en"):
        super().__init__()
        self.notebook_id = notebook_id
        self.title       = title
        self.source_id   = source_id
        self.lang        = lang

    def run(self):
        try:
            from notebooklm import NotebookLMClient
            async def _work():
                async with await NotebookLMClient.from_storage() as client:
                    # source_id có thể là "" (tất cả), 1 ID, hoặc nhiều ID phân tách bằng ","
                    if self.source_id:
                        parts = [s.strip() for s in self.source_id.split(",") if s.strip()]
                        sid = parts if parts else None
                    else:
                        sid = None
                    result = await client.chat.ask(self.notebook_id, _build_mindmap_prompt(self.lang),
                                                   source_ids=sid)
                    raw_text = (getattr(result, "answer", None) or str(result)).strip()
                    import json, re
                    m = re.search(r'\{.*\}', raw_text, re.DOTALL)
                    tree = json.loads(m.group(0)) if m else {"name": self.title, "children": []}
                    return _mindmap_to_html(tree, self.title)
            html = _run_async(_work())
            self.done.emit(html)
        except Exception as e:
            self.error.emit(str(e))


class NLMSummarizeWorker(QThread):
    """Upload file → chat.ask(summarize) → delete source."""
    done  = Signal(str)
    error = Signal(str)

    def __init__(self, notebook_id: str, file_path: str):
        super().__init__()
        self.notebook_id = notebook_id
        self.file_path   = file_path

    def run(self):
        try:
            from notebooklm import NotebookLMClient
            async def _work():
                async with await NotebookLMClient.from_storage() as client:
                    source = await client.sources.add_file(
                        self.notebook_id, Path(self.file_path), wait=True
                    )
                    sid = getattr(source, "source_id", None) or getattr(source, "id", None)
                    result = await client.chat.ask(
                        self.notebook_id,
                        _SUMMARIZE_PROMPT,
                        source_ids=[sid],
                    )
                    await client.sources.delete(self.notebook_id, sid)
                    return getattr(result, "answer", None) or str(result)
            text = _run_async(_work())
            self.done.emit(text)
        except Exception as e:
            self.error.emit(str(e))



class ChatWorker(QThread):
    done  = Signal(str, list)   # answer, deduplicated citations
    error = Signal(str)

    def __init__(self, notebook_id: str, question: str):
        super().__init__()
        self.notebook_id = notebook_id
        self.question    = question

    def run(self):
        try:
            from notebooklm import NotebookLMClient
            async def _chat():
                async with await NotebookLMClient.from_storage() as client:
                    result = await client.chat.ask(self.notebook_id, self.question)
                    # Build source_id → title map
                    src_map = {}
                    try:
                        sources = await client.sources.list(self.notebook_id)
                        for s in sources:
                            sid   = getattr(s, "source_id", None) or getattr(s, "id", "")
                            title = getattr(s, "title", None) or getattr(s, "name", "") or ""
                            if sid:
                                src_map[sid] = title
                    except Exception:
                        pass
                    return result, src_map
            result, src_map = _run_async(_chat())
            text = getattr(result, "answer", None) or getattr(result, "message", None) or getattr(result, "text", None) or str(result)

            seen, citations = set(), []
            for c in (getattr(result, "references", None) or []):
                quote  = (getattr(c, "cited_text", None) or "").strip()
                src_id = getattr(c, "source_id", "") or ""
                if quote and quote not in seen:
                    seen.add(quote)
                    citations.append({"text": quote, "source": src_map.get(src_id, "")})

            self.done.emit(text, citations)
        except Exception as e:
            self.error.emit(str(e))


class MultiChatWorker(QThread):
    """Query nhiều notebook song song, hiển thị từng cái có nội dung liên quan."""
    done  = Signal(list)   # [(nb_title, answer_text, citations_list), ...]
    error = Signal(str)

    def __init__(self, notebooks: list, question: str):
        super().__init__()
        self.notebooks = notebooks  # [(nb_id, nb_title), ...]
        self.question  = question

    def run(self):
        try:
            import asyncio
            from notebooklm import NotebookLMClient

            async def _ask_one(nb_id: str, nb_title: str):
                try:
                    async with await NotebookLMClient.from_storage() as client:
                        result = await client.chat.ask(nb_id, self.question)
                        src_map = {}
                        try:
                            sources = await client.sources.list(nb_id)
                            for s in sources:
                                sid   = getattr(s, "source_id", None) or getattr(s, "id", "")
                                title = getattr(s, "title", None) or getattr(s, "name", "") or ""
                                if sid:
                                    src_map[sid] = title
                        except Exception:
                            pass
                        text = (getattr(result, "answer", None)
                                or getattr(result, "message", None)
                                or getattr(result, "text", None)
                                or str(result))
                        seen, citations = set(), []
                        for c in (getattr(result, "references", None) or []):
                            quote  = (getattr(c, "cited_text", None) or "").strip()
                            src_id = getattr(c, "source_id", "") or ""
                            if quote and quote not in seen:
                                seen.add(quote)
                                citations.append({"text": quote, "source": src_map.get(src_id, "")})
                        return nb_title, text, citations
                except Exception as e:
                    return nb_title, f"⚠ Lỗi: {e}", []

            async def _run_all():
                tasks = [_ask_one(nb_id, nb_title) for nb_id, nb_title in self.notebooks]
                return await asyncio.gather(*tasks)

            raw_results = list(_run_async(_run_all()))
            self.done.emit(raw_results)
        except Exception as e:
            self.error.emit(str(e))


# ── Lọc nhiễu (bước 1) + gộp bằng Claude (bước 2) cho MultiChat ───────────

_EMPTY_ANSWER_HINTS = (
    "do not contain", "does not contain", "doesn't contain", "no information",
    "not contain information", "unable to find", "could not find", "couldn't find",
    "no relevant information", "no mention", "cannot find", "can't find",
    "do not provide", "does not provide", "not provide any", "no details",
    "không có thông tin", "không đề cập", "không tìm thấy", "nguồn không",
    "không chứa thông tin", "không nói", "chưa có thông tin", "không cung cấp",
    "không có chi tiết", "không có đủ thông tin", "không nêu", "không mô tả",
)


def _is_empty_answer(text: str) -> bool:
    """True nếu câu trả lời thực chất là 'không có thông tin liên quan' (kể cả khi dài dòng)."""
    t = (text or "").strip().lower()
    if not t:
        return True
    # Mở đầu bằng cụm phủ định → câu KHÔNG trả lời được, dù sau đó có giải thích dài
    head = t[:180]
    if any(h in head for h in _EMPTY_ANSWER_HINTS):
        return True
    # Câu ngắn có cụm phủ định ở bất kỳ đâu
    if len(t) < 400 and any(h in t for h in _EMPTY_ANSWER_HINTS):
        return True
    return False


def _build_reduce_prompt(question: str, items: list) -> str:
    """Prompt gộp các câu trả lời notebook thành 1, kèm cổng đánh giá liên quan."""
    parts = [
        f'Câu hỏi của người dùng:\n"{question}"\n',
        "Dưới đây là câu trả lời RIÊNG từ từng notebook. Hãy gộp thành MỘT câu trả lời "
        "cô đọng, chỉ giữ nội dung THỰC SỰ trả lời câu hỏi:",
    ]
    for i, (title, answer, _c) in enumerate(items, 1):
        parts.append(f"\n--- 📓 Notebook {i}: {title} ---\n{answer}\n")
    return "\n".join(parts)


class MultiChatReduceWorker(QThread):
    """Gộp N câu trả lời notebook thành 1 câu cô đọng bằng Claude (claude_agent_sdk)."""
    done  = Signal(str)
    error = Signal(str)

    def __init__(self, question: str, items: list):
        super().__init__()
        self.question = question
        self.items    = items   # [(nb_title, answer, citations), ...]

    def run(self):
        try:
            from claude_agent_sdk import query, ClaudeAgentOptions
            from claude_agent_sdk.types import AssistantMessage
            from paths import APP_DIR

            system = (
                "Bạn là trợ lý tổng hợp. Gộp câu trả lời từ nhiều notebook thành MỘT câu "
                "trả lời cô đọng, mạch lạc, bằng tiếng Việt. Yêu cầu:\n"
                "- Đánh giá mức liên quan của từng notebook với CÂU HỎI; LOẠI BỎ hẳn "
                "notebook lạc đề hoặc không trả lời được câu hỏi.\n"
                "- Hợp nhất các ý trùng lặp, không lặp lại.\n"
                "- Khi nêu thông tin quan trọng, ghi nguồn gọn dạng [📓 tên notebook].\n"
                "- Kết thúc bằng 1 dòng nhỏ: 'Nguồn: <các notebook đã dùng>' và nếu có "
                "loại bỏ thì thêm '(bỏ qua: <notebook> — không liên quan)'.\n"
                "- KHÔNG bịa thông tin ngoài các câu trả lời được cung cấp."
            )
            options = ClaudeAgentOptions(
                system_prompt=system,
                allowed_tools=[],
                permission_mode="bypassPermissions",
                cwd=APP_DIR,
                max_turns=1,
            )
            prompt = _build_reduce_prompt(self.question, self.items)

            async def _collect():
                buf = []
                async for message in query(prompt=prompt, options=options):
                    if isinstance(message, AssistantMessage):
                        for block in message.content:
                            if hasattr(block, "text") and block.text:
                                buf.append(block.text)
                return "".join(buf)

            text = _run_async(_collect())
            if not (text or "").strip():
                self.error.emit("empty")
                return
            self.done.emit(text)
        except Exception as e:
            self.error.emit(str(e))


class ImageChatWorker(QThread):
    """Vision AI mô tả ảnh dán → NbLM chat.ask với prompt đã enrich."""
    done  = Signal(str, list)
    error = Signal(str)

    _IMG_VISION_PROMPT = (
        "Describe this image in full detail. "
        "If it contains text, transcribe it verbatim. "
        "If it is a diagram, chart, table, or technical drawing, describe ALL elements, "
        "labels, values, connections, and relationships precisely. "
        "Be thorough — the description will be used as context for a Q&A system."
    )

    def __init__(self, notebook_id: str, question: str, image_bytes: bytes):
        super().__init__()
        self.notebook_id = notebook_id
        self.question    = question
        self.image_bytes = image_bytes

    def run(self):
        try:
            import base64
            from core.llm_config import load_llm_config

            cfg     = load_llm_config()
            img_b64 = base64.b64encode(self.image_bytes).decode()

            groq_key       = (cfg.get("groq_api_key")       or "").strip()
            gemini_key     = (cfg.get("gemini_api_key")     or "").strip()
            openrouter_key = (cfg.get("openrouter_api_key") or "").strip()
            gemini_model   = (cfg.get("gemini_model")       or "gemini-2.0-flash").strip()

            providers = []
            if groq_key:
                providers.append(lambda b: AddSourceWorker._call_openai_compat_vision(
                    base_url="https://api.groq.com/openai/v1",
                    api_key=groq_key,
                    model="meta-llama/llama-4-scout-17b-16e-instruct",
                    prompt=self._IMG_VISION_PROMPT, img_b64=b,
                ))
            if gemini_key:
                providers.append(lambda b, m=gemini_model: AddSourceWorker._call_gemini_vision(
                    api_key=gemini_key, model=m,
                    prompt=self._IMG_VISION_PROMPT, img_b64=b,
                ))
            if openrouter_key:
                providers.append(lambda b: AddSourceWorker._call_openai_compat_vision(
                    base_url="https://openrouter.ai/api/v1",
                    api_key=openrouter_key,
                    model="meta-llama/llama-4-scout:free",
                    prompt=self._IMG_VISION_PROMPT, img_b64=b,
                ))

            if not providers:
                raise RuntimeError(
                    "Chưa có API key Vision AI.\n"
                    "Vào ⚙ Config LLM để nhập key Groq / Gemini / OpenRouter."
                )

            vision_desc = ""
            last_err    = ""
            for fn in providers:
                try:
                    vision_desc = fn(img_b64)
                    if vision_desc:
                        break
                except Exception as e:
                    last_err = str(e)

            if not vision_desc:
                raise RuntimeError(f"Vision AI thất bại ở tất cả provider. Lỗi cuối: {last_err}")

            enriched = (
                f"[Hình ảnh người dùng đính kèm — mô tả tự động bởi Vision AI]\n"
                f"{vision_desc}\n\n"
                f"[Câu hỏi của người dùng]\n"
                f"{self.question}"
            )

            from notebooklm import NotebookLMClient
            async def _chat():
                async with await NotebookLMClient.from_storage() as client:
                    result = await client.chat.ask(self.notebook_id, enriched)
                    src_map = {}
                    try:
                        sources = await client.sources.list(self.notebook_id)
                        for s in sources:
                            sid   = getattr(s, "source_id", None) or getattr(s, "id", "")
                            title = getattr(s, "title", None) or getattr(s, "name", "") or ""
                            if sid:
                                src_map[sid] = title
                    except Exception:
                        pass
                    return result, src_map

            result, src_map = _run_async(_chat())
            text = (getattr(result, "answer", None) or getattr(result, "message", None)
                    or getattr(result, "text", None) or str(result))

            seen, citations = set(), []
            for c in (getattr(result, "references", None) or []):
                quote  = (getattr(c, "cited_text", None) or "").strip()
                src_id = getattr(c, "source_id", "") or ""
                if quote and quote not in seen:
                    seen.add(quote)
                    citations.append({"text": quote, "source": src_map.get(src_id, "")})

            self.done.emit(text, citations)
        except Exception as e:
            self.error.emit(str(e))


_TABLE_FORMAT_HINT = (
    "\n\n[Yêu cầu định dạng: Nếu dữ liệu phù hợp trình bày dạng bảng, "
    "hãy dùng markdown table với ký tự | phân cột và dòng --- phân header. "
    "Không cần giải thích thêm về định dạng, chỉ xuất bảng trực tiếp.]"
)

_TABLE_KEYWORDS = {
    "bảng", "table", "liệt kê", "list", "thống kê", "statistic",
    "danh sách", "so sánh", "compare", "tổng hợp", "summary",
    "thông số", "parameter", "dữ liệu", "data", "danh mục", "catalog",
    "chi tiết", "detail", "enumerate", "specification", "spec",
}


def _needs_table_hint(question: str) -> bool:
    q = question.lower()
    return any(kw in q for kw in _TABLE_KEYWORDS)


_SUGGEST_PROMPT = (
    "Based on the previous answer, generate exactly 8 concise follow-up questions "
    "that the user might want to ask next to deepen their understanding. "
    "Return ONLY a numbered list (1. ... 2. ... etc.), no introduction, no other text. "
    "Write questions in the same language as the previous answer."
)


class SuggestQuestionsWorker(QThread):
    """Ask NbLM to suggest follow-up questions based on the last answer."""
    done  = Signal(list)   # list[str] of question strings
    error = Signal(str)

    def __init__(self, notebook_id: str, last_answer: str):
        super().__init__()
        self.notebook_id = notebook_id
        self.last_answer = last_answer

    def run(self):
        try:
            import re
            from notebooklm import NotebookLMClient
            prompt = f"Previous answer:\n{self.last_answer[:2000]}\n\n{_SUGGEST_PROMPT}"
            async def _ask():
                async with await NotebookLMClient.from_storage() as client:
                    result = await client.chat.ask(self.notebook_id, prompt)
                    return getattr(result, "answer", None) or getattr(result, "message", None) or str(result)
            raw = _run_async(_ask())
            questions = []
            for line in raw.splitlines():
                line = line.strip()
                m = re.match(r"^\d+[\.\)]\s+(.+)$", line)
                if m:
                    q = m.group(1).strip()
                    if q and q not in questions:
                        questions.append(q)
            self.done.emit(questions[:10])
        except Exception as e:
            self.error.emit(str(e))


class SaveNoteWorker(QThread):
    """Save plain-text content as a note in the notebook."""
    done  = Signal(str)   # note id
    error = Signal(str)

    def __init__(self, notebook_id: str, title: str, content: str):
        super().__init__()
        self.notebook_id = notebook_id
        self.title   = title
        self.content = content

    def run(self):
        try:
            from notebooklm import NotebookLMClient
            async def _save():
                async with await NotebookLMClient.from_storage() as client:
                    note = await client.notes.create(self.notebook_id, self.title, self.content)
                    return getattr(note, "id", "") or ""
            self.done.emit(_run_async(_save()))
        except Exception as e:
            self.error.emit(str(e))


class ListNotesWorker(QThread):
    done  = Signal(list)
    error = Signal(str)

    def __init__(self, notebook_id: str):
        super().__init__()
        self.notebook_id = notebook_id

    def run(self):
        try:
            from notebooklm import NotebookLMClient
            async def _list():
                async with await NotebookLMClient.from_storage() as client:
                    return await client.notes.list(self.notebook_id)
            self.done.emit(_run_async(_list()))
        except Exception as e:
            self.error.emit(str(e))


class ConvertNoteToSourceWorker(QThread):
    done  = Signal(str)   # source id
    error = Signal(str)

    def __init__(self, notebook_id: str, note_id: str, title: str, content: str):
        super().__init__()
        self.notebook_id = notebook_id
        self.note_id  = note_id
        self.title    = title
        self.content  = content

    def run(self):
        try:
            from notebooklm import NotebookLMClient
            async def _conv():
                async with await NotebookLMClient.from_storage() as client:
                    src = await client.sources.add_text(
                        self.notebook_id, self.title, self.content)
                    return getattr(src, "id", "") or getattr(src, "source_id", "") or ""
            self.done.emit(_run_async(_conv()))
        except Exception as e:
            self.error.emit(str(e))


class GetLanguageWorker(QThread):
    done  = Signal(str)   # language code, e.g. "vi"
    error = Signal(str)

    def run(self):
        try:
            from notebooklm import NotebookLMClient
            async def _get():
                async with await NotebookLMClient.from_storage() as client:
                    return await client.settings.get_output_language() or ""
            self.done.emit(_run_async(_get()))
        except Exception as e:
            self.error.emit(str(e))


class SetLanguageWorker(QThread):
    done  = Signal(str)   # language code that was set
    error = Signal(str)

    def __init__(self, code: str):
        super().__init__()
        self.code = code

    def run(self):
        try:
            from notebooklm import NotebookLMClient
            from notebooklm.cli.language import set_language
            async def _set():
                async with await NotebookLMClient.from_storage() as client:
                    return await client.settings.set_output_language(self.code) or self.code
            result = _run_async(_set())
            self.done.emit(result)
        except Exception as e:
            self.error.emit(str(e))


# ── Studio Workers ────────────────────────────────────────────────────────────

class ListArtifactsWorker(QThread):
    done  = Signal(list)
    error = Signal(str)

    def __init__(self, notebook_id: str):
        super().__init__()
        self.notebook_id = notebook_id

    def run(self):
        try:
            from notebooklm import NotebookLMClient
            async def _list():
                async with await NotebookLMClient.from_storage() as client:
                    return await client.artifacts.list(self.notebook_id)
            self.done.emit(_run_async(_list()))
        except Exception as e:
            self.error.emit(str(e))


class GenerateArtifactWorker(QThread):
    """Generic worker: calls generate_fn(), polls until done, emits artifact id."""
    done    = Signal(str)   # artifact / task id
    status  = Signal(str)   # progress message
    error   = Signal(str)

    def __init__(self, notebook_id: str, kind: str, lang: str = "en",
                 report_format: str = "briefing_doc", instructions: str = "",
                 source_ids: list | None = None):
        super().__init__()
        self.notebook_id    = notebook_id
        self.kind           = kind
        self.lang           = lang
        self.report_format  = report_format
        self.instructions   = instructions
        self.source_ids     = source_ids or None  # None = dùng tất cả

    def run(self):
        try:
            from notebooklm import NotebookLMClient
            async def _gen():
                async with await NotebookLMClient.from_storage() as client:
                    k   = self.kind
                    sid = self.source_ids  # None = all sources
                    if k == "mind_map":
                        result = await client.artifacts.generate_mind_map(
                            self.notebook_id, source_ids=sid)
                        return result.get("note_id", "") or ""
                    elif k in ("briefing_doc", "study_guide", "blog_post"):
                        from notebooklm.rpc.types import ReportFormat
                        fmt_map = {
                            "briefing_doc": ReportFormat.BRIEFING_DOC,
                            "study_guide":  ReportFormat.STUDY_GUIDE,
                            "blog_post":    ReportFormat.BLOG_POST,
                        }
                        st = await client.artifacts.generate_report(
                            self.notebook_id, report_format=fmt_map[k],
                            language=self.lang,
                            extra_instructions=self.instructions or None,
                            source_ids=sid,
                        )
                    elif k == "flashcards":
                        st = await client.artifacts.generate_flashcards(
                            self.notebook_id,
                            instructions=self.instructions or None,
                            source_ids=sid,
                        )
                    elif k == "quiz":
                        st = await client.artifacts.generate_quiz(
                            self.notebook_id,
                            instructions=self.instructions or None,
                            source_ids=sid,
                        )
                    elif k == "data_table":
                        st = await client.artifacts.generate_data_table(
                            self.notebook_id, language=self.lang,
                            instructions=self.instructions or None,
                            source_ids=sid,
                        )
                    elif k == "infographic":
                        st = await client.artifacts.generate_infographic(
                            self.notebook_id, language=self.lang,
                            instructions=self.instructions or None,
                            source_ids=sid,
                        )
                    elif k == "slide_deck":
                        st = await client.artifacts.generate_slide_deck(
                            self.notebook_id, language=self.lang,
                            instructions=self.instructions or None,
                            source_ids=sid,
                        )
                    elif k == "audio":
                        st = await client.artifacts.generate_audio(
                            self.notebook_id, language=self.lang,
                            instructions=self.instructions or None,
                            source_ids=sid,
                        )
                    else:
                        raise ValueError(f"Unknown artifact kind: {k}")
                    try:
                        final = await client.artifacts.wait_for_completion(
                            self.notebook_id, st.task_id, timeout=600
                        )
                        return final.task_id or st.task_id
                    except TimeoutError:
                        # For media artifacts the server may still complete after
                        # our poll window. Return the task_id so the artifact
                        # list can be refreshed and the user can open it.
                        return st.task_id
            self.done.emit(_run_async(_gen()))
        except Exception as e:
            self.error.emit(str(e))


class OpenArtifactWorker(QThread):
    """Download artifact content; emits (kind, data) where data is str or bytes path."""
    done  = Signal(str, str)   # kind, content_or_path
    error = Signal(str)

    def __init__(self, notebook_id: str, artifact_id: str, kind: str):
        super().__init__()
        self.notebook_id = notebook_id
        self.artifact_id = artifact_id
        self.kind        = kind

    def run(self):
        import os, tempfile
        _NUM_KIND = {
            "1": "audio", "2": "report", "3": "video",
            "4": "quiz",  "5": "mind_map", "6": "flashcards",
            "7": "infographic", "8": "slide_deck", "9": "data_table",
        }
        try:
            from notebooklm import NotebookLMClient
            async def _open():
                async with await NotebookLMClient.from_storage() as client:
                    k   = _NUM_KIND.get(self.kind, self.kind)
                    tmp = tempfile.mkdtemp(prefix="finder_studio_")
                    if k == "mind_map":
                        # Download saved mind map JSON directly from NbLM (no re-generation)
                        import json as _json
                        path = os.path.join(tmp, "mind_map.json")
                        await client.artifacts.download_mind_map(
                            self.notebook_id, path, artifact_id=self.artifact_id)
                        with open(path, encoding="utf-8") as f:
                            tree = _json.load(f)
                        return ("mind_map_saved", _json.dumps(tree, ensure_ascii=False))
                    elif k in ("briefing_doc", "study_guide", "blog_post", "report"):
                        path = os.path.join(tmp, "report.md")
                        await client.artifacts.download_report(
                            self.notebook_id, path, artifact_id=self.artifact_id)
                        with open(path, encoding="utf-8") as f:
                            return ("report_md", f.read())
                    elif k == "flashcards":
                        path = os.path.join(tmp, "flashcards.html")
                        await client.artifacts.download_flashcards(
                            self.notebook_id, path,
                            artifact_id=self.artifact_id, output_format="html")
                        with open(path, encoding="utf-8") as f:
                            return ("html_view", f.read())
                    elif k == "quiz":
                        path = os.path.join(tmp, "quiz.html")
                        await client.artifacts.download_quiz(
                            self.notebook_id, path,
                            artifact_id=self.artifact_id, output_format="html")
                        with open(path, encoding="utf-8") as f:
                            return ("html_view", f.read())
                    elif k == "data_table":
                        path = os.path.join(tmp, "data_table.csv")
                        await client.artifacts.download_data_table(
                            self.notebook_id, path, artifact_id=self.artifact_id)
                        return ("csv_path", path)
                    elif k == "infographic":
                        path = os.path.join(tmp, "infographic.png")
                        await client.artifacts.download_infographic(
                            self.notebook_id, path, artifact_id=self.artifact_id)
                        return ("image_path", path)
                    elif k == "slide_deck":
                        path = os.path.join(tmp, "slides.pdf")
                        await client.artifacts.download_slide_deck(
                            self.notebook_id, path, artifact_id=self.artifact_id)
                        return ("file_path", path)
                    elif k == "audio":
                        path = os.path.join(tmp, "audio.mp4")
                        await client.artifacts.download_audio(
                            self.notebook_id, path, artifact_id=self.artifact_id)
                        return ("file_path", path)
                    elif k == "video":
                        path = os.path.join(tmp, "video.mp4")
                        await client.artifacts.download_video(
                            self.notebook_id, path, artifact_id=self.artifact_id)
                        return ("file_path", path)
                    else:
                        raise ValueError(f"Unknown kind: {k}")
            kind, data = _run_async(_open())
            self.done.emit(kind, data)
        except Exception as e:
            self.error.emit(str(e))


# ── End Studio Workers ────────────────────────────────────────────────────────


class PageFindWorker(QThread):
    """Runs _find_page_for_citation in background; emits found(page) when done."""
    found = Signal(int)   # page index (0-based)

    def __init__(self, pdf_path: str, cited_text: str):
        super().__init__()
        self.pdf_path   = pdf_path
        self.cited_text = cited_text

    def run(self):
        import re
        try:
            import fitz
            doc = fitz.open(self.pdf_path)
            try:
                # Pass 1: exact snippet (shrinking)
                for length in (60, 40, 20):
                    snippet = self.cited_text[:length].strip()
                    if len(snippet) < 8:
                        continue
                    for i in range(len(doc)):
                        if doc[i].search_for(snippet):
                            self.found.emit(i)
                            return
                # Pass 2: uppercase keyword matching
                keywords = list(dict.fromkeys(
                    re.findall(r'\b[A-Z][A-Z0-9_\-]{3,}\b', self.cited_text)))[:8]
                if keywords:
                    best_page, best_score = None, 0
                    for i in range(len(doc)):
                        score = sum(1 for k in keywords if k in doc[i].get_text())
                        if score > best_score:
                            best_score, best_page = score, i
                    if best_score >= 3:
                        self.found.emit(best_page)
            finally:
                doc.close()
        except Exception:
            pass
        # Not found — do not emit (preview stays at page 0)


# ── Notebook Picker Dialog (used from external windows) ──────────

class NLMNotebookPickerDialog(QDialog):
    """Show list of notebooks, return selected notebook_id."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Select Notebook")
        self.setFixedSize(360, 280)
        self.selected_id = None

        lay = QVBoxLayout(self)
        lay.addWidget(QLabel("Select a notebook to summarize into:"))

        self.lst = QListWidget()
        lay.addWidget(self.lst, 1)

        btns = QHBoxLayout()
        ok_btn = QPushButton("OK")
        ok_btn.setFixedHeight(32)
        ok_btn.clicked.connect(self._accept)
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setFixedHeight(32)
        cancel_btn.clicked.connect(self.reject)
        btns.addStretch()
        btns.addWidget(ok_btn)
        btns.addWidget(cancel_btn)
        lay.addLayout(btns)

        self._load()

    def _load(self):
        self.lst.addItem(QListWidgetItem("⏳ Loading…"))
        w = ListNotebooksWorker()
        w.done.connect(self._on_loaded)
        w.error.connect(lambda e: (self.lst.clear(), self.lst.addItem(f"Error: {e}")))
        self._worker = w
        w.start()

    def _on_loaded(self, notebooks):
        self.lst.clear()
        for nb in notebooks:
            title = getattr(nb, "title", None) or getattr(nb, "name", str(nb))
            nb_id = getattr(nb, "id", None) or getattr(nb, "notebook_id", "")
            item  = QListWidgetItem(f"📓 {title}")
            item.setData(Qt.UserRole, nb_id)
            self.lst.addItem(item)
        if self.lst.count():
            self.lst.setCurrentRow(0)

    def _accept(self):
        item = self.lst.currentItem()
        if item:
            self.selected_id = item.data(Qt.UserRole)
            self.accept()

    @staticmethod
    def pick(parent=None) -> str | None:
        if not is_nlm_logged_in():
            QMessageBox.warning(
                parent, "Not Logged In",
                "Please log in to NotebookLM first.\n\nGo to the 📓 NotebookLM tab and click 🔑 Login Google."
            )
            return None
        dlg = NLMNotebookPickerDialog(parent)
        if dlg.exec() == QDialog.Accepted:
            return dlg.selected_id
        return None


# ── Image paste event filter ─────────────────────────────────────

class _ImagePasteFilter(QObject):
    """Intercept Ctrl+V trên QLineEdit khi clipboard chứa ảnh."""
    image_pasted = Signal(object)   # QImage

    def eventFilter(self, obj, event):
        from PySide6.QtCore import QEvent
        from PySide6.QtGui import QKeySequence
        if event.type() == QEvent.Type.KeyPress:
            if event.matches(QKeySequence.StandardKey.Paste):
                from PySide6.QtWidgets import QApplication
                mime = QApplication.clipboard().mimeData()
                if mime.hasImage():
                    img = QApplication.clipboard().image()
                    if not img.isNull():
                        self.image_pasted.emit(img)
                        return True   # consume — không paste text
        return False


# ── Chat display with clickable source links ─────────────────────

class ChatDisplay(QTextEdit):
    link_clicked     = Signal(str)
    question_clicked = Signal(str)   # suggested question text

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setReadOnly(True)
        self._suggest_questions: list[str] = []

    def set_suggestions(self, questions: list[str]):
        self._suggest_questions = questions

    def mousePressEvent(self, e):
        anchor = self.anchorAt(e.pos())
        if anchor and anchor.startswith("nlm://"):
            self.link_clicked.emit(anchor)
        elif anchor and anchor.startswith("xlsx://"):
            import os
            path = anchor[7:]
            if os.path.isfile(path):
                os.startfile(path)
        elif anchor and anchor.startswith("q://"):
            try:
                idx = int(anchor[4:])
                q = self._suggest_questions[idx]
                self.question_clicked.emit(q)
            except Exception:
                pass
        else:
            super().mousePressEvent(e)


# ── Embedded Widget ───────────────────────────────────────────────

class NotebookLMWidget(QWidget):
    """Embedded widget — dùng trong tab hoặc dialog."""
    open_preview         = Signal(str, object)  # file_path, page_num (int or None)
    goto_page_signal     = Signal(int)          # jump to page after background search
    request_mindmap        = Signal(str)          # file_path → tạo mind map từ file local
    request_mindmap_online = Signal(str, str, str)  # notebook_id, source_id, title → tạo mind map online

    def __init__(self, parent=None):
        super().__init__(parent)
        self._citation_refs: list = []
        self._current_language: str = "en"
        self._last_answer: str = ""

        self._current_notebook_id = None
        self._checked_notebook_ids: dict[str, str] = {}  # {nb_id: title}
        self._pasted_image_bytes: bytes | None = None
        self._workers = []
        self._thinking_step = 0
        self._thinking_timer = QTimer(self)
        self._thinking_timer.setInterval(500)
        self._thinking_timer.timeout.connect(self._tick_thinking)

        self._setup_ui()

        # Auto-load nếu đã có session
        if is_nlm_logged_in():
            self.lbl_status.setText("🟢 Logged in")
            self._load_notebooks()
            self._fetch_current_language()
        else:
            self.lbl_status.setText("⚪ Not logged in")

    def _start_worker(self, w):
        """Append worker, start it, and auto-remove when finished."""
        self._workers.append(w)
        w.finished.connect(lambda: self._workers.remove(w) if w in self._workers else None)
        w.finished.connect(w.deleteLater)
        w.start()
        return w

    def _setup_ui(self):
        lay = QVBoxLayout(self)
        lay.setSpacing(8)
        lay.setContentsMargins(8, 8, 8, 8)

        # ── Header: Login + status ────────────────────────────────
        header = QHBoxLayout()
        self.lbl_status = QLabel("⚪ Not logged in")
        self.lbl_status.setStyleSheet("font-size: 12px; color: #a6adc8;")
        self.lbl_status.setMaximumWidth(320)
        self.btn_login = QPushButton("🔑 Switch Account")
        self.btn_login.setFixedHeight(32)
        self.btn_login.clicked.connect(self._login)
        self.btn_save_login = QPushButton("✅ Save Login")
        self.btn_save_login.setFixedHeight(32)
        self.btn_save_login.setEnabled(False)
        self.btn_save_login.setToolTip("Click after you have finished logging in on the browser")
        self.btn_save_login.clicked.connect(self._save_login)
        self.btn_language = QPushButton("🌐 …")
        self.btn_language.setFixedHeight(32)
        self.btn_language.setToolTip("Set NbLM output language (global setting)")
        self.btn_language.clicked.connect(self._pick_language)
        self.btn_llm_config = QPushButton("⚙ Config LLM")
        self.btn_llm_config.setFixedHeight(32)
        self.btn_llm_config.setToolTip("Config LLM models & API keys")
        self.btn_llm_config.clicked.connect(self._open_llm_settings)
        self.btn_refresh = QPushButton("🔄 Refresh")
        self.btn_refresh.setFixedHeight(32)
        self.btn_refresh.clicked.connect(self._load_notebooks)
        self.btn_toggle_studio = QPushButton("🎨 Studio")
        self.btn_toggle_studio.setFixedHeight(32)
        self.btn_toggle_studio.setToolTip("Toggle Studio panel")
        self.btn_toggle_studio.setCheckable(True)
        self.btn_toggle_studio.clicked.connect(self._toggle_studio_panel)
        header.addWidget(self.lbl_status)
        header.addStretch()
        header.addWidget(self.btn_login)
        header.addWidget(self.btn_save_login)
        header.addWidget(self.btn_language)
        header.addWidget(self.btn_llm_config)
        header.addWidget(self.btn_refresh)
        header.addWidget(self.btn_toggle_studio)
        lay.addLayout(header)

        # ── Splitter: notebooks list | main panel ─────────────────
        splitter = QSplitter(Qt.Horizontal)

        # Left: notebook list
        left = QWidget()
        left_lay = QVBoxLayout(left)
        left_lay.setContentsMargins(0, 0, 0, 0)
        left_lay.setSpacing(6)

        nb_header = QHBoxLayout()
        lbl_nb = QLabel("📓 Notebooks")
        lbl_nb.setStyleSheet("font-weight: bold; font-size: 13px;")
        nb_header.addWidget(lbl_nb)
        nb_header.addStretch()
        self.btn_check_all_nb = QPushButton("☑ Check all")
        self.btn_check_all_nb.setFixedHeight(30)
        self.btn_check_all_nb.setToolTip("Chọn / bỏ chọn tất cả sổ")
        self.btn_check_all_nb.setEnabled(False)
        self.btn_check_all_nb.setStyleSheet("""
            QPushButton {
                background: #1e3a5f; color: #93c5fd;
                border: 1px solid #3b82f6; border-radius: 6px;
                font-size: 12px; padding: 0 10px;
            }
            QPushButton:hover   { background: #2a4f80; color: #bfdbfe; }
            QPushButton:pressed { background: #1a3050; }
            QPushButton:disabled { background: #1a2535; color: #4a6080; border-color: #2a3a4a; }
        """)
        self.btn_check_all_nb.clicked.connect(self._toggle_check_all_notebooks)
        nb_header.addWidget(self.btn_check_all_nb)
        self.cmb_nb_sort = QComboBox()
        self.cmb_nb_sort.addItem("A→Z",  "alpha_asc")
        self.cmb_nb_sort.addItem("Z→A",  "alpha_desc")
        self.cmb_nb_sort.addItem("Newest", "date_desc")
        self.cmb_nb_sort.addItem("Oldest", "date_asc")
        self.cmb_nb_sort.setFixedHeight(26)
        self.cmb_nb_sort.setFixedWidth(80)
        self.cmb_nb_sort.setToolTip("Sort notebooks")
        self.cmb_nb_sort.currentIndexChanged.connect(self._resort_notebooks)
        nb_header.addWidget(self.cmb_nb_sort)
        left_lay.addLayout(nb_header)

        self.lst_notebooks = QTreeWidget()
        self.lst_notebooks.setHeaderHidden(True)
        self.lst_notebooks.setIndentation(14)
        self.lst_notebooks.setStyleSheet("""
            QTreeWidget { font-size: 15px; }
            QTreeWidget::item { padding: 4px 2px; }
            QTreeWidget::item:selected { background: #313244; }
        """)
        self.lst_notebooks.setContextMenuPolicy(Qt.CustomContextMenu)
        self.lst_notebooks.customContextMenuRequested.connect(self._nb_context_menu)
        self.lst_notebooks.currentItemChanged.connect(self._on_notebook_selected)
        self.lst_notebooks.itemExpanded.connect(self._on_notebook_tree_expanded)
        self.lst_notebooks.itemChanged.connect(self._on_tree_item_check_changed)
        left_lay.addWidget(self.lst_notebooks, 1)

        nb_btn_row = QHBoxLayout()
        nb_btn_row.setSpacing(4)
        self.btn_new_nb = QPushButton("➕ New")
        self.btn_new_nb.setFixedHeight(34)
        self.btn_new_nb.clicked.connect(self._create_notebook)
        self.btn_del_nb = QPushButton("🗑 Delete")
        self.btn_del_nb.setFixedHeight(34)
        self.btn_del_nb.setEnabled(False)
        self.btn_del_nb.clicked.connect(self._delete_notebook)
        self.btn_merge_nb = QPushButton("⊕ Merge")
        self.btn_merge_nb.setFixedHeight(34)
        self.btn_merge_nb.setEnabled(False)
        self.btn_merge_nb.setToolTip(
            "Gộp nội dung các notebook đã check thành 1 notebook mới.\n"
            "Nội dung lấy từ DB đã index — không cần re-upload file."
        )
        self.btn_merge_nb.clicked.connect(self._merge_notebooks)
        nb_btn_row.addWidget(self.btn_new_nb)
        nb_btn_row.addWidget(self.btn_del_nb)
        nb_btn_row.addWidget(self.btn_merge_nb)
        left_lay.addLayout(nb_btn_row)



        splitter.addWidget(left)

        # Right: tabs
        right = QTabWidget()
        self._right_tabs = right

        # Tab 1: Chat
        chat_widget = QWidget()
        chat_lay = QVBoxLayout(chat_widget)
        chat_lay.setSpacing(6)

        self.lbl_nb_name = QLabel("Select a notebook →")
        self.lbl_nb_name.setStyleSheet("font-size: 13px; font-weight: bold; color: #89b4fa;")
        chat_lay.addWidget(self.lbl_nb_name)

        self.chat_display = ChatDisplay()
        self.chat_display.setPlaceholderText("Chat history will appear here…")
        self.chat_display.setStyleSheet("font-size: 18px;")
        self.chat_display.link_clicked.connect(self._on_source_link_clicked)
        self.chat_display.question_clicked.connect(self._on_suggestion_clicked)
        chat_lay.addWidget(self.chat_display, 1)

        self.lbl_thinking = QLabel("")
        self.lbl_thinking.setStyleSheet("font-size: 14px; color: #89b4fa; padding: 2px 4px;")
        self.lbl_thinking.setVisible(False)
        chat_lay.addWidget(self.lbl_thinking)

        # ── Image indicator row (ẩn mặc định) ──────────────────
        img_row = QHBoxLayout()
        self.lbl_img_thumb = QLabel()
        self.lbl_img_thumb.setFixedSize(36, 36)
        self.lbl_img_thumb.setScaledContents(True)
        self.lbl_img_thumb.setStyleSheet(
            "border:1px solid #3b82f6;border-radius:4px;background:#0f172a;")
        self.lbl_img_thumb.setVisible(False)
        self.lbl_img_indicator = QLabel("📷 Image attached")
        self.lbl_img_indicator.setStyleSheet(
            "color:#93c5fd;font-size:12px;padding:0 4px;")
        self.lbl_img_indicator.setVisible(False)
        self.btn_clear_img = QPushButton("✕")
        self.btn_clear_img.setFixedSize(20, 20)
        self.btn_clear_img.setToolTip("Remove attached image")
        self.btn_clear_img.setVisible(False)
        self.btn_clear_img.clicked.connect(self._clear_pasted_image)
        img_row.addWidget(self.lbl_img_thumb)
        img_row.addWidget(self.lbl_img_indicator)
        img_row.addWidget(self.btn_clear_img)
        img_row.addStretch()
        chat_lay.addLayout(img_row)

        input_row = QHBoxLayout()
        self.chat_input = QLineEdit()
        self.chat_input.setPlaceholderText("Ask a question about your documents… (Ctrl+V để dán ảnh)")
        self.chat_input.setFixedHeight(36)
        self.chat_input.returnPressed.connect(self._send_chat)
        # Install image paste event filter
        self._img_filter = _ImagePasteFilter(self)
        self._img_filter.image_pasted.connect(self._on_image_pasted)
        self.chat_input.installEventFilter(self._img_filter)
        self.btn_send = QPushButton("Send ➤")
        self.btn_send.setFixedHeight(36)
        self.btn_send.setEnabled(False)
        self.btn_send.clicked.connect(self._send_chat)
        self.btn_save_note = QPushButton("📌 Note")
        self.btn_save_note.setFixedHeight(36)
        self.btn_save_note.setEnabled(False)
        self.btn_save_note.setToolTip("Save this chat response as a note")
        self.btn_save_note.clicked.connect(self._save_chat_as_note)
        input_row.addWidget(self.chat_input, 1)
        input_row.addWidget(self.btn_send)
        input_row.addWidget(self.btn_save_note)
        chat_lay.addLayout(input_row)

        right.addTab(chat_widget, "💬 Chat")

        # Tab 2: Sources
        src_widget = QWidget()
        src_lay = QVBoxLayout(src_widget)
        src_lay.setSpacing(6)

        # Header row: label + Select All checkbox
        src_header = QHBoxLayout()
        src_header.addWidget(QLabel("Files added to this notebook:"))
        src_header.addStretch()
        self.chk_select_all_src = QPushButton("☐ Select All")
        self.chk_select_all_src.setFlat(True)
        self.chk_select_all_src.setFixedHeight(24)
        self.chk_select_all_src.setEnabled(False)
        self.chk_select_all_src.clicked.connect(self._toggle_select_all_sources)
        src_header.addWidget(self.chk_select_all_src)
        src_lay.addLayout(src_header)

        self.lst_sources = QListWidget()
        self.lst_sources.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.lst_sources.setTextElideMode(Qt.ElideNone)
        self.lst_sources.setContextMenuPolicy(Qt.CustomContextMenu)
        self.lst_sources.customContextMenuRequested.connect(self._show_source_context_menu)
        self.lst_sources.itemChanged.connect(self._on_source_check_changed)
        src_lay.addWidget(self.lst_sources, 1)

        # Delete action row has been removed here
        src_btn_row = QHBoxLayout()
        self.btn_add_file = QPushButton("📄 Add File")
        self.btn_add_file.setFixedHeight(30)
        self.btn_add_file.setEnabled(False)
        self.btn_add_file.clicked.connect(self._add_file)
        
        from PySide6.QtWidgets import QCheckBox
        self.chk_vision = QCheckBox("📡 Diagram Analysis (Vision AI)")
        self.chk_vision.setToolTip(
            "Use Vision AI to understand logic in technical diagrams (FBD, P&ID, wiring…)."
            "\nRequires an API key (Groq/Gemini/OpenRouter) and Internet. Only applies to PDFs."
        )

        self.btn_add_folder = QPushButton("📁 Folder")
        self.btn_add_folder.setFixedHeight(30)
        self.btn_add_folder.setEnabled(True)
        self.btn_add_folder.setToolTip(
            "Chọn folder → tự tạo notebook mới đặt tên theo folder.\n"
            "Cứ 100 file sẽ tự tạo thêm notebook mới (Part 2, 3…)."
        )
        self.btn_add_folder.clicked.connect(self._add_folder)

        src_btn_row.addWidget(self.btn_add_file)
        src_btn_row.addWidget(self.btn_add_folder)
        src_btn_row.addWidget(self.chk_vision)
        src_btn_row.addStretch()
        src_lay.addLayout(src_btn_row)

        self.lbl_batch_progress = QLabel("")
        self.lbl_batch_progress.setStyleSheet("font-size:12px; color:#89b4fa; padding:2px 0;")
        self.lbl_batch_progress.setVisible(False)
        src_lay.addWidget(self.lbl_batch_progress)

        right.addTab(src_widget, "📎 Sources")

        # ── Studio tab ──────────────────────────────────────────
        studio_widget = QWidget()
        studio_lay    = QVBoxLayout(studio_widget)
        studio_lay.setSpacing(8)

        from PySide6.QtWidgets import QGroupBox, QGridLayout
        gen_group = QGroupBox("✨ Generate")
        gen_grid  = QGridLayout(gen_group)
        gen_grid.setSpacing(6)

        _STUDIO_BTNS = [
            ("🗺 Mind Map",     "mind_map"),
            ("📄 Briefing Doc", "briefing_doc"),
            ("📚 Study Guide",  "study_guide"),
            ("🃏 Flashcards",   "flashcards"),
            ("❓ Quiz",         "quiz"),
            ("📊 Data Table",   "data_table"),
            ("🖼 Infographic",  "infographic"),
            ("🎞 Slide Deck",   "slide_deck"),
            ("🎙 Audio",        "audio"),
        ]
        self._studio_gen_btns = {}
        for idx, (label, kind) in enumerate(_STUDIO_BTNS):
            btn = QPushButton(label)
            btn.setFixedHeight(40)
            btn.setEnabled(False)
            btn.clicked.connect(lambda _=False, k=kind: self._studio_generate(k))
            gen_grid.addWidget(btn, idx // 3, idx % 3)
            self._studio_gen_btns[kind] = btn
        studio_lay.addWidget(gen_group)

        self.lbl_studio_status = QLabel("")
        self.lbl_studio_status.setWordWrap(True)
        studio_lay.addWidget(self.lbl_studio_status)

        from PySide6.QtWidgets import QSplitter as _QSplitter
        studio_inner = _QSplitter(Qt.Vertical)

        # Notes section
        notes_widget = QWidget()
        notes_lay = QVBoxLayout(notes_widget)
        notes_lay.setContentsMargins(0, 0, 0, 0)
        notes_lay.setSpacing(4)
        notes_lay.addWidget(QLabel("📝 Notes:"))
        self.lst_notes = QListWidget()
        self.lst_notes.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.lst_notes.setTextElideMode(Qt.ElideNone)
        self.lst_notes.setContextMenuPolicy(Qt.CustomContextMenu)
        self.lst_notes.customContextMenuRequested.connect(self._show_note_context_menu)
        self.lst_notes.itemDoubleClicked.connect(self._open_note_item)
        notes_lay.addWidget(self.lst_notes, 1)
        studio_inner.addWidget(notes_widget)

        # Artifacts section
        arts_widget = QWidget()
        arts_lay = QVBoxLayout(arts_widget)
        arts_lay.setContentsMargins(0, 0, 0, 0)
        arts_lay.setSpacing(4)
        arts_lay.addWidget(QLabel("📁 Existing artifacts:"))
        self.lst_artifacts = QListWidget()
        self.lst_artifacts.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.lst_artifacts.setTextElideMode(Qt.ElideNone)
        self.lst_artifacts.setContextMenuPolicy(Qt.CustomContextMenu)
        self.lst_artifacts.customContextMenuRequested.connect(
            self._show_artifact_context_menu)
        self.lst_artifacts.itemDoubleClicked.connect(self._open_artifact_item)
        arts_lay.addWidget(self.lst_artifacts, 1)

        art_btn_row = QHBoxLayout()
        btn_refresh_arts = QPushButton("🔄 Refresh")
        btn_refresh_arts.setFixedHeight(30)
        btn_refresh_arts.clicked.connect(self._load_studio)
        art_btn_row.addWidget(btn_refresh_arts)
        art_btn_row.addStretch()
        arts_lay.addLayout(art_btn_row)
        studio_inner.addWidget(arts_widget)

        studio_inner.setSizes([200, 300])
        studio_lay.addWidget(studio_inner, 1)

        right.currentChanged.connect(self._on_right_tab_changed)

        # Inner splitter: Chat+Sources tabs | Studio panel
        inner_splitter = QSplitter(Qt.Horizontal)
        inner_splitter.addWidget(right)
        inner_splitter.addWidget(studio_widget)
        inner_splitter.setStretchFactor(0, 1)
        inner_splitter.setStretchFactor(1, 0)
        # Studio panel hidden by default
        studio_widget.setVisible(False)
        self._inner_splitter  = inner_splitter
        self._studio_widget   = studio_widget

        splitter.addWidget(inner_splitter)
        splitter.setSizes([240, 760])

        lay.addWidget(splitter, 1)

    # ── Auth ─────────────────────────────────────────────────────

    def _login(self):
        self.btn_login.setEnabled(False)
        self.lbl_status.setText("🔄 Opening browser for login…")
        self._login_worker = LoginWorker()
        self._login_worker.done.connect(self._on_login_done)
        self._login_worker.error.connect(self._on_login_error)
        self._workers.append(self._login_worker)
        self._login_worker.start()
        self.btn_save_login.setEnabled(True)

    def _save_login(self):
        if hasattr(self, "_login_worker"):
            self._login_worker.confirm()
        self.btn_save_login.setEnabled(False)
        self.lbl_status.setText("🔄 Saving login…")

    def _on_login_done(self):
        self.lbl_status.setText("🟢 Logged in")
        self.btn_login.setEnabled(True)
        self._load_notebooks()
        self._fetch_current_language()

    def _fetch_current_language(self):
        w = GetLanguageWorker()
        w.done.connect(self._on_language_fetched)
        w.error.connect(lambda _: None)
        w.finished.connect(w.deleteLater)
        self._start_worker(w)

    def _on_language_fetched(self, code: str):
        self._current_language = code or "en"
        self.btn_language.setText(f"🌐 {self._current_language}")

    def _pick_language(self):
        from PySide6.QtWidgets import QDialog, QVBoxLayout, QComboBox, QLabel, QDialogButtonBox
        LANGUAGES = {
            "en": "English",
            "vi": "Tiếng Việt",
            "ko": "한국어",
            "ja": "日本語",
        }
        dlg = QDialog(self)
        dlg.setWindowTitle("🌐 Output Language")
        dlg.setFixedWidth(280)
        lay = QVBoxLayout(dlg)
        lay.addWidget(QLabel("NbLM output language\n(global — affects all notebooks)"))
        combo = QComboBox()
        codes = list(LANGUAGES.keys())
        for code, name in LANGUAGES.items():
            combo.addItem(f"{name}  ({code})", code)
        # Pre-select current language from button text
        cur = self.btn_language.text().replace("🌐 ", "").strip()
        if cur in codes:
            combo.setCurrentIndex(codes.index(cur))
        lay.addWidget(combo)
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)
        if dlg.exec() != QDialog.Accepted:
            return
        code = combo.currentData()
        self.btn_language.setText(f"🌐 ⏳")
        self.btn_language.setEnabled(False)
        w = SetLanguageWorker(code)
        w.done.connect(lambda c: (
            self._on_language_fetched(c),
            self.btn_language.setEnabled(True),
        ))
        w.error.connect(lambda e: (
            QMessageBox.critical(self, "Error", e),
            self.btn_language.setEnabled(True),
        ))
        w.finished.connect(w.deleteLater)
        self._start_worker(w)

    def _open_llm_settings(self):
        from ui.pdf_preview import LLMSettingsDialog
        dlg = LLMSettingsDialog(self)
        dlg.exec()

    def _on_login_error(self, msg: str):
        self.lbl_status.setText("🔴 Login failed")
        self.btn_login.setEnabled(True)
        QMessageBox.critical(self, "Login Error", msg)

    # ── Notebooks ─────────────────────────────────────────────────

    def _load_notebooks(self):
        self.lbl_status.setText("🔄 Loading notebooks…")
        w = ListNotebooksWorker()
        w.done.connect(self._on_notebooks_loaded)
        w.error.connect(lambda e: (
            self.lbl_status.setText("🔴 Auth expired — click Switch Account"),
            self.lbl_status.setToolTip(e),
        ))
        self._start_worker(w)

    @staticmethod
    def _notebook_emoji(title: str) -> str:
        t = title.lower()
        if any(k in t for k in ("boiler", "lo hoi", "steam", "furnace", "burner")): return "🔥"
        if any(k in t for k in ("turbine", "generator", "rotor")): return "⚙️"
        if any(k in t for k in ("electric", "dien", "điện", "power")): return "⚡"
        if any(k in t for k in ("pump", "bom", "bơm", "valve", "van")): return "🔧"
        if any(k in t for k in ("safety", "an toan", "an toàn", "hazard", "alarm")): return "⚠️"
        if any(k in t for k in ("chemistry", "water", "nuoc", "nước", "chemical")): return "💧"
        if any(k in t for k in ("procedure", "sop", "operation", "quy trinh", "quy trình")): return "📋"
        if any(k in t for k in ("drawing", "ban ve", "bản vẽ", "piping", "diagram")): return "📐"
        if any(k in t for k in ("report", "bao cao", "báo cáo", "summary")): return "📊"
        if any(k in t for k in ("training", "huan luyen", "huấn luyện", "course")): return "🎓"
        return "📓"

    def _on_notebooks_loaded(self, notebooks):
        self._notebooks_raw = list(notebooks)
        self._render_notebooks(self._notebooks_raw)
        self.lbl_status.setText(f"🟢 {len(notebooks)} notebook(s) loaded")

    def _sort_key(self, nb, mode: str):
        if mode in ("alpha_asc", "alpha_desc"):
            return (getattr(nb, "title", None) or "").lower()
        else:
            dt = getattr(nb, "created_at", None)
            import datetime
            return dt if dt else datetime.datetime.min

    def _resort_notebooks(self):
        if not hasattr(self, "_notebooks_raw"):
            return
        mode = self.cmb_nb_sort.currentData()
        notebooks = sorted(
            self._notebooks_raw,
            key=lambda nb: self._sort_key(nb, mode),
            reverse=(mode in ("alpha_desc", "date_desc")),
        )
        self._render_notebooks(notebooks)

    def _render_notebooks(self, notebooks):
        self.lst_notebooks.clear()
        self._checked_notebook_ids.clear()
        self.btn_check_all_nb.setEnabled(bool(notebooks))
        self.btn_check_all_nb.setText("☑ Check all")
        for nb in notebooks:
            title = getattr(nb, "title", None) or getattr(nb, "name", str(nb))
            nb_id = getattr(nb, "id", None) or getattr(nb, "notebook_id", "")
            emoji = self._notebook_emoji(title)
            item  = QTreeWidgetItem([f"{emoji}  {title}"])
            item.setData(0, Qt.UserRole, nb_id)
            item.setData(0, Qt.UserRole + 1, "notebook")
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(0, Qt.Unchecked)
            font = item.font(0)
            font.setBold(True)
            item.setFont(0, font)
            placeholder = QTreeWidgetItem(["⏳"])
            placeholder.setData(0, Qt.UserRole + 1, "placeholder")
            item.addChild(placeholder)
            self.lst_notebooks.addTopLevelItem(item)

    def _create_notebook(self):
        from PySide6.QtWidgets import QInputDialog
        title, ok = QInputDialog.getText(self, "New Notebook", "Notebook title:")
        if not ok or not title.strip():
            return
        w = CreateNotebookWorker(title.strip())
        w.done.connect(lambda nb: self._load_notebooks())
        w.error.connect(lambda e: QMessageBox.critical(self, "Error", e))
        self._start_worker(w)

    def _nb_context_menu(self, pos):
        from PySide6.QtWidgets import QMenu
        item = self.lst_notebooks.itemAt(pos)
        if not item:
            return
        kind = item.data(0, Qt.UserRole + 1) or ""
        if kind == "source":
            # Source child context menu
            menu = QMenu(self)
            menu.addAction("👁 View Content").triggered.connect(
                lambda: self._view_source_content(item))
            menu.addAction("🗺 Mind Map").triggered.connect(
                lambda: self._mindmap_from_source(item))
            menu.addSeparator()
            menu.addAction("🗑 Delete Source").triggered.connect(
                lambda: self._delete_source(item))
            menu.exec(self.lst_notebooks.mapToGlobal(pos))
            return
        if kind != "notebook":
            return
        menu = QMenu(self)
        act_rename = menu.addAction("✏️ Rename")
        act_delete = menu.addAction("🗑 Delete")
        action = menu.exec(self.lst_notebooks.mapToGlobal(pos))
        if action == act_rename:
            self._rename_notebook(item)
        elif action == act_delete:
            self._delete_notebook()

    def _rename_notebook(self, item):
        from PySide6.QtWidgets import QInputDialog
        nb_id = item.data(0, Qt.UserRole)
        old_title = item.text(0).split("  ", 1)[-1]  # bỏ emoji prefix
        new_title, ok = QInputDialog.getText(
            self, "Rename Notebook", "New name:", text=old_title
        )
        if not ok or not new_title.strip() or new_title.strip() == old_title:
            return
        w = RenameNotebookWorker(nb_id, new_title.strip())
        w.done.connect(lambda t: self._load_notebooks())
        w.error.connect(lambda e: QMessageBox.critical(self, "Error", e))
        self._start_worker(w)

    def _delete_notebook(self):
        item = self.lst_notebooks.currentItem()
        if not item or item.data(0, Qt.UserRole + 1) != "notebook":
            return
        title = item.text(0)
        nb_id = item.data(0, Qt.UserRole)
        reply = QMessageBox.question(
            self, "Delete Notebook",
            f"Delete \"{title}\"?\nThis cannot be undone.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        self.btn_del_nb.setEnabled(False)
        w = DeleteNotebookWorker(nb_id)
        w.done.connect(lambda: self._load_notebooks())
        w.error.connect(lambda e: QMessageBox.critical(self, "Error", e))
        self._start_worker(w)

    def _on_notebook_selected(self, current, _prev):
        if not current:
            self.btn_del_nb.setEnabled(False)
            return
        kind = current.data(0, Qt.UserRole + 1) or ""
        if kind == "source":
            return
        if kind != "notebook":
            return
        self._current_notebook_id = current.data(0, Qt.UserRole)
        title = current.text(0).split("  ", 1)[-1]
        self.lbl_nb_name.setText(f"📓 {title}")
        self.btn_del_nb.setEnabled(True)
        self.btn_send.setEnabled(True)
        self.btn_save_note.setEnabled(True)
        self.btn_add_file.setEnabled(True)
        self._set_studio_btns_enabled(True)
        self.chat_display.clear()
        self._load_sources()

    def _toggle_check_all_notebooks(self):
        """Check All / Uncheck All cho toàn bộ notebook trong list."""
        root = self.lst_notebooks.invisibleRootItem()
        count = root.childCount()
        if count == 0:
            return
        all_checked = all(
            root.child(i).checkState(0) == Qt.Checked
            for i in range(count)
        )
        new_state = Qt.Unchecked if all_checked else Qt.Checked
        self.lst_notebooks.blockSignals(True)
        self._checked_notebook_ids.clear()
        for i in range(count):
            item = root.child(i)
            item.setCheckState(0, new_state)
            if new_state == Qt.Checked:
                nb_id    = item.data(0, Qt.UserRole) or ""
                nb_title = item.text(0)
                if nb_id:
                    self._checked_notebook_ids[nb_id] = nb_title
        self.lst_notebooks.blockSignals(False)
        self.btn_check_all_nb.setText("☐ Uncheck all" if new_state == Qt.Checked else "☑ Check all")
        if self._checked_notebook_ids and not self._current_notebook_id:
            self.btn_send.setEnabled(True)
        self.btn_merge_nb.setEnabled(len(self._checked_notebook_ids) >= 2)

    def _on_tree_item_check_changed(self, item, col):
        """Khi tick/untick item trong tree → xử lý notebook multi-select, Select All, sync sources."""
        if col != 0:
            return
        kind = item.data(0, Qt.UserRole + 1) or ""
        parent = item.parent()

        if kind == "notebook":
            nb_id    = item.data(0, Qt.UserRole) or ""
            nb_title = item.text(0)
            if item.checkState(0) == Qt.Checked:
                self._checked_notebook_ids[nb_id] = nb_title
            else:
                self._checked_notebook_ids.pop(nb_id, None)
            # Enable Send nếu có ít nhất 1 sổ được check (kể cả chưa click sổ nào)
            if self._checked_notebook_ids and not self._current_notebook_id:
                self.btn_send.setEnabled(True)
            self.btn_merge_nb.setEnabled(len(self._checked_notebook_ids) >= 2)
            return

        self.lst_notebooks.blockSignals(True)
        if kind == "select_all" and parent:
            # Toggle tất cả source con theo trạng thái Select All
            state = item.checkState(0)
            item.setText(0, "☑ Select All" if state == Qt.Checked else "☐ Select All")
            for i in range(parent.childCount()):
                child = parent.child(i)
                if child.data(0, Qt.UserRole + 1) == "source":
                    child.setCheckState(0, state)
        elif kind == "source" and parent:
            # Cập nhật lại dòng Select All của notebook cha
            total = sum(1 for i in range(parent.childCount())
                        if parent.child(i).data(0, Qt.UserRole + 1) == "source")
            checked = sum(1 for i in range(parent.childCount())
                          if parent.child(i).data(0, Qt.UserRole + 1) == "source"
                          and parent.child(i).checkState(0) == Qt.Checked)
            for i in range(parent.childCount()):
                ch = parent.child(i)
                if ch.data(0, Qt.UserRole + 1) == "select_all":
                    new_state = Qt.Checked if checked == total else Qt.Unchecked
                    ch.setCheckState(0, new_state)
                    ch.setText(0, "☑ Select All" if new_state == Qt.Checked else "☐ Select All")
                    break
        self.lst_notebooks.blockSignals(False)

        if kind != "source":
            return
        # Sync sang lst_sources
        sid   = item.data(0, Qt.UserRole)
        state = item.checkState(0)
        self.lst_sources.blockSignals(True)
        for i in range(self.lst_sources.count()):
            it = self.lst_sources.item(i)
            if it and it.data(Qt.UserRole) == sid:
                it.setCheckState(state)
                break
        self.lst_sources.blockSignals(False)
        self._update_src_action_bar()

    def _on_notebook_tree_expanded(self, item):
        """Load sources as children when notebook is expanded."""
        if item.data(0, Qt.UserRole + 1) != "notebook":
            return
        nb_id = item.data(0, Qt.UserRole)
        # Check if already loaded (not placeholder)
        if item.childCount() == 1:
            child = item.child(0)
            if child.data(0, Qt.UserRole + 1) == "placeholder":
                child.setText(0, "⏳ Loading…")
                self._load_sources_into_tree(nb_id, item)

    def _load_sources_into_tree(self, nb_id: str, parent_item):
        w = ListSourcesWorker(nb_id)
        w.done.connect(lambda sources: self._on_tree_sources_loaded(sources, parent_item))
        w.error.connect(lambda e: (
            parent_item.child(0).setText(0, f"Error: {e}") if parent_item.childCount() else None
        ))
        self._start_worker(w)

    def _on_tree_sources_loaded(self, sources, parent_item):
        # Ghi nhớ trạng thái checked trước khi xóa children (để giữ khi reload)
        prev_checked: set[str] = set()
        was_first_load = True
        for i in range(parent_item.childCount()):
            ch = parent_item.child(i)
            if ch.data(0, Qt.UserRole + 1) == "source":
                was_first_load = False
                if ch.checkState(0) == Qt.Checked:
                    prev_checked.add(ch.data(0, Qt.UserRole))

        while parent_item.childCount():
            parent_item.removeChild(parent_item.child(0))

        if not sources:
            empty = QTreeWidgetItem(["(no sources)"])
            empty.setData(0, Qt.UserRole + 1, "empty")
            parent_item.addChild(empty)
            return

        self.lst_notebooks.blockSignals(True)
        # Dòng Select All
        sel_all = QTreeWidgetItem(["☑ Select All"])
        sel_all.setData(0, Qt.UserRole + 1, "select_all")
        sel_all.setFlags(sel_all.flags() | Qt.ItemIsUserCheckable)
        sel_all.setCheckState(0, Qt.Checked)
        parent_item.addChild(sel_all)

        all_checked = True
        for s in sources:
            title     = getattr(s, "title", None) or str(s)
            sid       = getattr(s, "source_id", None) or getattr(s, "id", "") or ""
            type_code = getattr(s, "_type_code", None)
            icon, _   = self._SOURCE_TYPE_GROUPS.get(type_code, ("📁", "Other"))
            # Lần đầu load → check hết; reload → giữ trạng thái cũ
            state = Qt.Checked if (was_first_load or sid in prev_checked) else Qt.Unchecked
            if state == Qt.Unchecked:
                all_checked = False
            child = QTreeWidgetItem([f"{icon} {title}"])
            child.setData(0, Qt.UserRole,     sid)
            child.setData(0, Qt.UserRole + 1, "source")
            child.setToolTip(0, title)
            child.setFlags(child.flags() | Qt.ItemIsUserCheckable)
            child.setCheckState(0, state)
            parent_item.addChild(child)

        # Cập nhật dòng Select All cho khớp
        if not all_checked:
            sel_all.setCheckState(0, Qt.Unchecked)
            sel_all.setText(0, "☐ Select All")
        self.lst_notebooks.blockSignals(False)

    def _reload_current_notebook_sources(self):
        """Reload sources in the tree for the currently selected notebook."""
        if not self._current_notebook_id:
            return
        for i in range(self.lst_notebooks.topLevelItemCount()):
            item = self.lst_notebooks.topLevelItem(i)
            if item.data(0, Qt.UserRole) == self._current_notebook_id:
                if item.isExpanded():
                    self._load_sources_into_tree(self._current_notebook_id, item)
                break
        self._load_sources()   # also refresh Sources tab

    def _load_sources(self):
        if not self._current_notebook_id:
            return
        self.lst_sources.clear()
        self.lst_sources.addItem("⏳ Loading…")
        w = ListSourcesWorker(self._current_notebook_id)
        w.done.connect(self._on_sources_loaded)
        w.error.connect(lambda e: (self.lst_sources.clear(), self.lst_sources.addItem(f"Error: {e}")))
        self._start_worker(w)

    # Source type code → (icon, group label)
    _SOURCE_TYPE_GROUPS = {
        1: ("📊", "Google Docs"),
        2: ("📊", "Google Drive"),
        3: ("📄", "PDF"),
        4: ("📝", "Pasted Text"),
        5: ("🌐", "Web Page"),
        8: ("✨", "Generated"),
        9: ("🎥", "YouTube"),
    }

    def _on_sources_loaded(self, sources):
        self.lst_sources.clear()
        if not sources:
            self.lst_sources.addItem("(no sources)")
            return
        self.lst_sources.blockSignals(True)
        for s in sources:
            title = getattr(s, "title", None) or getattr(s, "name", None) or str(s)
            sid   = getattr(s, "source_id", None) or getattr(s, "id", None) or ""
            item  = QListWidgetItem(f"📄 {title}")
            item.setData(Qt.UserRole, sid)
            item.setToolTip(title)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Unchecked)
            self.lst_sources.addItem(item)
        self.lst_sources.blockSignals(False)
        self.chk_select_all_src.setEnabled(bool(sources))
        self._update_src_action_bar()

    def _show_source_context_menu(self, pos):
        item = self.lst_sources.itemAt(pos)
        if not item or not item.data(Qt.UserRole):
            return   # group header — no menu
        from PySide6.QtWidgets import QMenu
        menu = QMenu(self)
        menu.addAction("👁 View Content").triggered.connect(
            lambda: self._view_source_content(item)
        )
        menu.addAction("🗺 Create Mind Map").triggered.connect(
            lambda: self._mindmap_from_source(item)
        )
        menu.addSeparator()
        menu.addAction("🗑 Delete Source").triggered.connect(
            lambda: self._delete_source(item)
        )
        menu.exec(self.lst_sources.mapToGlobal(pos))

    def _mindmap_from_source(self, item):
        try:
            from PySide6.QtWidgets import QTreeWidgetItem as _QTWI
            if isinstance(item, _QTWI):
                title     = item.text(0).strip()
                source_id = item.data(0, Qt.UserRole) or ""
            else:
                title     = item.text().strip()
                source_id = item.data(Qt.UserRole) or ""
        except RuntimeError:
            return # item obj has been deleted by a background refresh
        local_path = _lookup_source_path(title)
        if not local_path:
            from PySide6.QtWidgets import QMessageBox
            reply = QMessageBox.question(self, "File Not Found",
                f'Local file not found for "{title}".\n\n'
                "Generate mind map online from this source?",
                QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                self._mindmap_online(title, source_id)
            return
        self.request_mindmap.emit(local_path)

    def _mindmap_online(self, source_title: str, source_id: str = ""):
        """Tạo mind map từ source đang chọn, hiển thị trong panel preview."""
        if not self._current_notebook_id:
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.warning(self, "No Notebook", "Please select a notebook first.")
            return
        self.request_mindmap_online.emit(self._current_notebook_id, source_id, source_title)

    def _delete_source(self, item):
        from PySide6.QtWidgets import QMessageBox
        try:
            title = item.text().strip()
            sid   = item.data(Qt.UserRole)
        except RuntimeError:
            return # item obj has been deleted by a background refresh
        if not sid:
            QMessageBox.warning(self, "Error", "Cannot delete: source ID not found.")
            return
        reply = QMessageBox.question(self, "Delete Source",
            f'Delete "{title}" from this notebook?',
            QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return
        w = DeleteSourceWorker(self._current_notebook_id, sid)
        w.done.connect(lambda: self._reload_current_notebook_sources())
        w.error.connect(lambda e: QMessageBox.critical(self, "Error", e))
        self._start_worker(w)

    def _view_source_content(self, item):
        """Fetch and display full text content of a source in the chat panel."""
        try:
            title     = item.text().strip()
            source_id = item.data(Qt.UserRole)
        except RuntimeError:
            return
        if not source_id or not self._current_notebook_id:
            return

        self._right_tabs.setCurrentIndex(0)  # switch to Chat tab
        self.chat_display.append(f"<b style='color:#89b4fa'>📄 {title}</b>")
        self.chat_display.append("<i style='color:#888'>⏳ Loading content…</i>")

        class _FetchWorker(QThread):
            done  = Signal(str, int)
            error = Signal(str)
            def __init__(self, nb_id, src_id):
                super().__init__()
                self.nb_id  = nb_id
                self.src_id = src_id
            def run(self):
                try:
                    from notebooklm import NotebookLMClient
                    async def _get():
                        async with await NotebookLMClient.from_storage() as client:
                            ft = await client.sources.get_fulltext(self.nb_id, self.src_id)
                            content = getattr(ft, "content", "") or ""
                            chars   = getattr(ft, "char_count", len(content))
                            return content, int(chars)
                    self.done.emit(*_run_async(_get()))
                except Exception as e:
                    self.error.emit(str(e))

        def _on_done(content, chars):
            # Remove the "⏳ Loading…" line
            cursor = self.chat_display.textCursor()
            from PySide6.QtGui import QTextCursor
            cursor.movePosition(QTextCursor.End)
            cursor.select(QTextCursor.BlockUnderCursor)
            cursor.removeSelectedText()
            cursor.deletePreviousChar()
            # Append content as plain preformatted block
            import html as _html
            escaped = _html.escape(content)
            self.chat_display.append(
                f"<span style='white-space:pre-wrap'>{escaped}</span>"
                f"<br><i style='color:#888'>— {chars:,} chars —</i><br>"
            )

        def _on_err(msg):
            self.chat_display.append(f"<i style='color:#f38ba8'>Error: {msg}</i><br>")

        fw = _FetchWorker(self._current_notebook_id, source_id)
        fw.done.connect(_on_done)
        fw.error.connect(_on_err)
        self._start_worker(fw)

    # ── Chat ──────────────────────────────────────────────────────

    def _start_thinking(self):
        self._thinking_step = 0
        self.lbl_thinking.setVisible(True)
        self._thinking_timer.start()

    def _stop_thinking(self):
        self._thinking_timer.stop()
        self.lbl_thinking.setVisible(False)

    def _tick_thinking(self):
        frames = ["🤔 Thinking", "🤔 Thinking·", "🤔 Thinking··", "🤔 Thinking···"]
        self.lbl_thinking.setText(frames[self._thinking_step % len(frames)])
        self._thinking_step += 1

    def select_notebook_for_file(self, file_path: str) -> bool:
        """Try to auto-select the notebook that contains file_path. Returns True if found."""
        src_map = _load_source_map()
        fname = os.path.basename(file_path)
        # Find entry by path or filename
        nb_id = None
        for entry in src_map.values():
            if not isinstance(entry, dict):
                continue
            if entry.get("path", "") == file_path or os.path.basename(entry.get("path", "")) == fname:
                nbs = entry.get("notebooks", [])
                if nbs:
                    nb_id = nbs[0]
                    break
        if not nb_id:
            return False
        # Find matching item in lst_notebooks and select it
        for i in range(self.lst_notebooks.topLevelItemCount()):
            item = self.lst_notebooks.topLevelItem(i)
            if item.data(0, Qt.UserRole) == nb_id:
                self.lst_notebooks.setCurrentItem(item)
                return True
        return False

    def open_temp_chat_notebook(self, nb_id: str, nb_title: str):
        """Select temp notebook in list (or add it if not present), focus chat input."""
        for i in range(self.lst_notebooks.topLevelItemCount()):
            item = self.lst_notebooks.topLevelItem(i)
            if item.data(0, Qt.UserRole) == nb_id:
                self.lst_notebooks.setCurrentItem(item)
                self._current_notebook_id = nb_id
                self.lbl_nb_name.setText(f"📓 {nb_title}")
                self.btn_send.setEnabled(True)
                self.btn_save_note.setEnabled(True)
                self._load_sources()
                self.chat_input.setFocus()
                return
        # Not in list yet — add it and select
        item = QTreeWidgetItem([f"📓 {nb_title}"])
        item.setData(0, Qt.UserRole,     nb_id)
        item.setData(0, Qt.UserRole + 1, "notebook")
        placeholder = QTreeWidgetItem(["⏳"])
        placeholder.setData(0, Qt.UserRole + 1, "placeholder")
        item.addChild(placeholder)
        font = item.font(0)
        font.setBold(True)
        item.setFont(0, font)
        self.lst_notebooks.addTopLevelItem(item)
        self.lst_notebooks.setCurrentItem(item)
        self._current_notebook_id = nb_id
        self.lbl_nb_name.setText(f"📓 {nb_title}")
        self.btn_del_nb.setEnabled(True)
        self.btn_send.setEnabled(True)
        self.btn_save_note.setEnabled(True)
        self.btn_add_file.setEnabled(True)
        self._set_studio_btns_enabled(True)
        self.chat_display.clear()
        self._load_sources()
        self.chat_input.setFocus()

    def ask_from_mindmap(self, node_name: str):
        """Called externally (e.g. from mind map node) to ask NbLM about a node."""
        if not self._current_notebook_id:
            return
        question = f'Explain "{node_name}" in detail based on the document.'
        self.chat_input.setText(question)
        self._send_chat()
        # Switch to Chat tab
        for i in range(self._right_tabs.count()):
            if self._right_tabs.tabText(i).startswith("💬"):
                self._right_tabs.setCurrentIndex(i)
                break

    def _on_image_pasted(self, qimage):
        """Lưu ảnh vừa dán và hiện thumbnail preview."""
        from PySide6.QtCore import QBuffer, QIODevice
        from PySide6.QtGui import QPixmap
        buf = QBuffer()
        buf.open(QIODevice.OpenMode.WriteOnly)
        qimage.save(buf, "PNG")
        self._pasted_image_bytes = bytes(buf.data())
        buf.close()
        pix = QPixmap.fromImage(qimage).scaled(
            36, 36, Qt.AspectRatioMode.KeepAspectRatio,
            Qt.TransformationMode.SmoothTransformation,
        )
        self.lbl_img_thumb.setPixmap(pix)
        self.lbl_img_thumb.setVisible(True)
        sz = qimage.size()
        self.lbl_img_indicator.setText(f"📷 {sz.width()}×{sz.height()} px")
        self.lbl_img_indicator.setVisible(True)
        self.btn_clear_img.setVisible(True)

    def _clear_pasted_image(self):
        """Xóa ảnh đã dán."""
        self._pasted_image_bytes = None
        self.lbl_img_thumb.clear()
        self.lbl_img_thumb.setVisible(False)
        self.lbl_img_indicator.setVisible(False)
        self.btn_clear_img.setVisible(False)

    def _send_chat(self):
        if not self._current_notebook_id and not self._checked_notebook_ids:
            return
        question = self.chat_input.text().strip()
        if not question and not self._pasted_image_bytes:
            return
        if not question:
            question = "Hãy mô tả và phân tích hình ảnh này."

        img_bytes = self._pasted_image_bytes
        self._clear_pasted_image()

        # Hiện ảnh trong chat nếu có
        if img_bytes:
            from PySide6.QtGui import QImage, QTextDocument
            from PySide6.QtCore import QUrl
            qimg = QImage.fromData(img_bytes)
            # Scale để không chiếm quá nhiều không gian
            if qimg.width() > 320 or qimg.height() > 320:
                qimg = qimg.scaled(
                    320, 320,
                    Qt.AspectRatioMode.KeepAspectRatio,
                    Qt.TransformationMode.SmoothTransformation,
                )
            img_name = f"pasted_{id(img_bytes)}"
            self.chat_display.document().addResource(
                QTextDocument.ResourceType.ImageResource,
                QUrl(img_name),
                qimg,
            )
            self.chat_display.append(
                f"<b style='color:#89b4fa'>You:</b> {question}<br>"
                f"<img src='{img_name}' width='{qimg.width()}' height='{qimg.height()}'>"
            )
        else:
            self.chat_display.append(f"<b style='color:#89b4fa'>You:</b> {question}")

        self.chat_input.clear()
        self.btn_send.setEnabled(False)
        self.btn_save_note.setEnabled(False)
        self._start_thinking()

        if self._checked_notebook_ids and not img_bytes:
            # Multi-notebook mode: query tất cả sổ đã check song song
            nb_list = list(self._checked_notebook_ids.items())  # [(id, title), ...]
            self._multi_question = question   # giữ câu hỏi gốc cho bước gộp
            sent_question = question + (_TABLE_FORMAT_HINT if _needs_table_hint(question) else "")
            w = MultiChatWorker(nb_list, sent_question)
            w.done.connect(self._on_multi_chat_done)
            w.error.connect(self._on_chat_error)
        elif img_bytes:
            w = ImageChatWorker(self._current_notebook_id, question, img_bytes)
            w.done.connect(self._on_chat_done)
            w.error.connect(self._on_chat_error)
        else:
            sent_question = question + (_TABLE_FORMAT_HINT if _needs_table_hint(question) else "")
            w = ChatWorker(self._current_notebook_id, sent_question)
            w.done.connect(self._on_chat_done)
            w.error.connect(self._on_chat_error)
        self._start_worker(w)

    def _on_chat_done(self, text: str, citations: list):
        self._last_answer = text
        self._stop_thinking()
        self._citation_refs = citations
        html = self._md_to_html(text)
        self.chat_display.append(f"<b style='color:#a6e3a1'>Mr Finder:</b><br>{html}")
        if citations:
            parts = []
            for i, c in enumerate(citations, 1):
                quote  = c.get("text", "")
                source = c.get("source", "")
                short  = quote[:150] + ("…" if len(quote) > 150 else "")
                src_html = (
                    f" <a href='nlm://{i-1}' style='color:#1d4ed8;text-decoration:underline'>{source}</a>"
                    if source else ""
                )
                parts.append(
                    f"<span style='color:#15803d'>[{i}]{src_html}</span>"
                    f" <i style='color:#374151'>\"{short}\"</i>"
                )
            self.chat_display.append(
                "<span style='font-size:15px;color:#b45309'>──────────────── Sources ────────────────</span><br>"
                + "<br>".join(parts)
            )
        self.chat_display.append("<br>")
        self.btn_send.setEnabled(True)
        self.btn_save_note.setEnabled(True)
        # Auto-export Excel if response contains a markdown table
        self._auto_export_excel_if_table(text)
        # Generate suggested follow-up questions in background
        self.chat_display.set_suggestions([])
        self.chat_display.append(
            "<span style='color:#6c7086;font-size:14px'>⏳ Suggested questions…</span>"
        )
        sw = SuggestQuestionsWorker(self._current_notebook_id, text)
        sw.done.connect(self._on_suggestions_done)
        sw.error.connect(lambda _: None)   # fail silently
        sw.finished.connect(sw.deleteLater)
        self._start_worker(sw)

    def _on_multi_chat_done(self, results: list):
        """Bước 1: lọc notebook trả lời thật. Bước 2: gộp bằng Claude → 1 câu cô đọng."""
        # Lọc: không lỗi + không phải câu 'không có thông tin'.
        # KHÔNG bắt buộc có citations — NotebookLM nhiều khi trả lời đúng mà không kèm
        # citation, trước đây bị bỏ luôn nên 'thinking xong mà trống trơn'.
        relevant = [
            r for r in results
            if not r[1].startswith("⚠") and not _is_empty_answer(r[1])
        ]

        if not relevant:
            self._stop_thinking()
            self._last_answer   = ""
            self._citation_refs = []
            self.chat_display.append(
                "<i style='color:#6c7086'>Không tìm thấy nội dung liên quan trong các "
                "notebook đã chọn.</i>"
            )
            self.chat_display.append("<br>")
            self.btn_send.setEnabled(True)
            self.btn_save_note.setEnabled(False)
            return

        # Bước 2: gộp bằng Claude (giữ spinner; xong ở _on_reduce_done)
        self._pending_multi = relevant
        rw = MultiChatReduceWorker(getattr(self, "_multi_question", ""), relevant)
        rw.done.connect(self._on_reduce_done)
        rw.error.connect(self._on_reduce_error)
        self._start_worker(rw)

    def _on_reduce_done(self, text: str):
        """Hiển thị câu trả lời tổng hợp + phần Sources gộp từ các notebook liên quan."""
        self._stop_thinking()
        self._last_answer = text
        html = self._md_to_html(text)
        self.chat_display.append(
            "<b style='color:#a6e3a1'>Mr Finder</b> "
            "<span style='color:#89b4fa'>[tổng hợp]</span>:<br>" + html
        )

        # Gộp citations từ các notebook liên quan (giữ tên notebook để biết nguồn)
        cites, seen = [], set()
        for nb_title, _ans, citations in (getattr(self, "_pending_multi", []) or []):
            for c in (citations or []):
                quote  = c.get("text", "")
                source = c.get("source", "")
                key = (quote, source)
                if quote and key not in seen:
                    seen.add(key)
                    cites.append({"text": quote, "source": source, "notebook": nb_title})
        self._citation_refs = cites

        if cites:
            parts = []
            for i, c in enumerate(cites, 1):
                quote  = c["text"]
                source = c["source"]
                nbt    = c.get("notebook", "")
                short  = quote[:150] + ("…" if len(quote) > 150 else "")
                src_html = (
                    f" <a href='nlm://{i-1}' style='color:#1d4ed8;text-decoration:underline'>{source}</a>"
                    if source else ""
                )
                nb_html = f" <span style='color:#89b4fa'>[📓 {nbt}]</span>" if nbt else ""
                parts.append(
                    f"<span style='color:#15803d'>[{i}]{src_html}</span>{nb_html}"
                    f" <i style='color:#374151'>\"{short}\"</i>"
                )
            self.chat_display.append(
                "<span style='font-size:15px;color:#b45309'>──────────────── Sources ────────────────</span><br>"
                + "<br>".join(parts)
            )

        self.chat_display.append("<br>")
        self.btn_send.setEnabled(True)
        self.btn_save_note.setEnabled(True)

    def _on_reduce_error(self, _msg: str):
        """Gộp lỗi (vd chưa đăng nhập Claude) → báo rõ rồi fallback hiển thị từng notebook."""
        self.chat_display.append(
            "<span style='color:#b45309;font-size:13px'>⚠ Chưa gộp được câu trả lời "
            "(cần đăng nhập Claude ở tab Claude). Tạm hiển thị riêng từng notebook:</span>"
        )
        self._render_multi_per_notebook(getattr(self, "_pending_multi", []) or [])

    def _render_multi_per_notebook(self, relevant: list):
        """Fallback: hiển thị riêng từng notebook + citations (luồng cũ)."""
        self._stop_thinking()
        if not relevant:
            self.btn_send.setEnabled(True)
            self.btn_save_note.setEnabled(False)
            return
        combined_text = []
        for nb_title, text, citations in relevant:
            combined_text.append(text)
            self.chat_display.append(
                f"<b style='color:#a6e3a1'>Mr Finder</b> <span style='color:#89b4fa'>[📓 {nb_title}]</span>:"
            )
            self.chat_display.append(self._md_to_html(text))
            if citations:
                parts = []
                for i, c in enumerate(citations, 1):
                    quote  = c.get("text", "")
                    source = c.get("source", "")
                    short  = quote[:150] + ("…" if len(quote) > 150 else "")
                    src_html = (
                        f" <a href='nlm://{i-1}' style='color:#1d4ed8;text-decoration:underline'>{source}</a>"
                        if source else ""
                    )
                    parts.append(
                        f"<span style='color:#15803d'>[{i}]{src_html}</span>"
                        f" <i style='color:#374151'>\"{short}\"</i>"
                    )
                self.chat_display.append(
                    "<span style='font-size:15px;color:#b45309'>──────────────── Sources ────────────────</span><br>"
                    + "<br>".join(parts)
                )
            self.chat_display.append("<br>")

        self._last_answer   = "\n\n---\n\n".join(combined_text)
        self._citation_refs = []
        self.btn_send.setEnabled(True)
        self.btn_save_note.setEnabled(True)

    def _on_suggestions_done(self, questions: list[str]):
        if not questions:
            return
        self.chat_display.set_suggestions(questions)
        # Remove the "⏳ Suggested questions…" placeholder
        doc = self.chat_display.document()
        found = doc.find("⏳ Suggested questions…")
        if not found.isNull():
            found.select(found.SelectionType.LineUnderCursor)
            found.removeSelectedText()
        # Append header + each question as a separate paragraph
        self.chat_display.append(
            "<span style='color:#6c7086;font-size:13px'>💡 Suggested questions:</span>"
        )
        for i, q in enumerate(questions):
            self.chat_display.append(
                f"<a href='q://{i}' style='color:#0000FF;text-decoration:none;font-size:16px'>"
                f"▸ {q}</a>"
            )
        self.chat_display.append("<br>")

    def _on_source_link_clicked(self, href: str):
        try:
            idx = int(href.replace("nlm://", ""))
            c = self._citation_refs[idx]
            source     = c.get("source", "")
            cited_text = c.get("text", "")
        except Exception:
            return

        # Tìm trong source map
        local_path = _lookup_source_path(source)

        if not local_path:
            from PySide6.QtWidgets import QFileDialog, QMessageBox
            answer = QMessageBox.question(self, "File not found",
                f'"{source}" not in local map.\nLocate the file manually?',
                QMessageBox.Yes | QMessageBox.No)
            if answer != QMessageBox.Yes:
                return
            local_path, _ = QFileDialog.getOpenFileName(
                self, f"Locate: {source}", "", "PDF Files (*.pdf)")
            if not local_path:
                return
            _upsert_source_map(source, local_path, self._current_notebook_id or "")

        if not local_path.lower().endswith(".pdf"):
            return
        # Open preview immediately (no lag), then find the right page in background
        self.open_preview.emit(local_path, None)
        if cited_text.strip():
            worker = PageFindWorker(local_path, cited_text)
            worker.found.connect(self.goto_page_signal)
            worker.finished.connect(worker.deleteLater)
            # Keep a reference so GC doesn't collect it before it finishes
            if not hasattr(self, "_page_find_workers"):
                self._page_find_workers = []
            self._page_find_workers.append(worker)
            worker.finished.connect(lambda: self._page_find_workers.remove(worker)
                                    if worker in self._page_find_workers else None)
            worker.start()

    _LATEX_MAP = {
        r"\rightarrow": "→", r"\leftarrow": "←",
        r"\Rightarrow": "⇒", r"\Leftarrow": "⇐",
        r"\leftrightarrow": "↔", r"\Leftrightarrow": "⇔",
        r"\leq": "≤", r"\geq": "≥", r"\neq": "≠",
        r"\approx": "≈", r"\times": "×", r"\div": "÷",
        r"\pm": "±", r"\infty": "∞", r"\cdot": "·",
        r"\alpha": "α", r"\beta": "β", r"\gamma": "γ",
        r"\delta": "δ", r"\Delta": "Δ", r"\mu": "μ",
        r"\pi": "π", r"\sigma": "σ", r"\theta": "θ",
        r"\sum": "∑", r"\prod": "∏", r"\int": "∫",
        r"\sqrt": "√", r"\partial": "∂",
    }

    @staticmethod
    def _replace_latex(text: str) -> str:
        import re
        def _sub(m):
            inner = m.group(1).strip()
            for sym, uni in NotebookLMWidget._LATEX_MAP.items():
                inner = inner.replace(sym, uni)
            # nếu còn lại chỉ là text thuần → trả về text không có $
            return inner
        return re.sub(r"\$([^$]+)\$", _sub, text)

    @staticmethod
    def _md_to_html(text: str) -> str:
        """Convert markdown to HTML using markdown-it-py with Pygments and Light Blue Theme."""
        try:
            from markdown_it import MarkdownIt
            from mdit_py_plugins.ext_table import ext_table # If not available, markdown-it handles basic tables via enable('table')
        except ImportError:
            pass

        text = NotebookLMWidget._replace_latex(text)
        
        try:
            md = MarkdownIt('commonmark', {'breaks': True, 'html': True}).enable('table')
            
            # Use Pygments for code highlighting if available
            try:
                from pygments import highlight
                from pygments.lexers import get_lexer_by_name, guess_lexer
                from pygments.formatters import HtmlFormatter
                
                def highlight_code(code: str, lang: str, attrs) -> str:
                    try:
                        lexer = get_lexer_by_name(lang, stripall=True) if lang else guess_lexer(code)
                    except Exception:
                        from pygments.lexers.special import TextLexer
                        lexer = TextLexer()
                    formatter = HtmlFormatter(style="github-dark", cssclass="pygments-code")
                    return highlight(code, lexer, formatter)
                
                md.options["highlight"] = highlight_code
                pygments_css = HtmlFormatter(style="github-dark", cssclass="pygments-code").get_style_defs()
            except ImportError:
                pygments_css = ""
                
            raw_html = md.render(text)
            
            # CSS specifically for Light Blue theme tables and nicely formatted blocks
            THEME_CSS = f"""
            <style>
                {pygments_css}
                table {{
                    border-collapse: collapse;
                    width: 100%;
                    margin: 12px 0;
                    font-size: 14px;
                    border: 1px solid #bae6fd;
                }}
                th {{
                    background-color: #0284c7;
                    color: white;
                    font-weight: bold;
                    padding: 8px;
                    border: 1px solid #7dd3fc;
                    text-align: left;
                }}
                td {{
                    padding: 8px;
                    border: 1px solid #e0f2fe;
                    color: #000000;
                    font-weight: bold;
                }}
                tr:nth-child(even) td {{
                    background-color: rgba(14, 165, 233, 0.1);
                }}
                pre {{
                    background-color: #111827;
                    padding: 10px;
                    border-radius: 6px;
                    border: 1px solid #374151;
                    overflow-x: auto;
                }}
                code {{
                    font-family: Consolas, monospace;
                    font-size: 13px;
                }}
                blockquote {{
                    border-left: 4px solid #0ea5e9;
                    padding-left: 12px;
                    color: #9ca3af;
                    margin-left: 0;
                }}
            </style>
            """
            return THEME_CSS + raw_html
        except Exception:
            # Fallback if markdown_it fails
            text = text.replace("<", "&lt;").replace(">", "&gt;")
            import re
            lines = text.split("\n")
            out = []
            for line in lines:
                if line.startswith("### "): out.append(f"<b style='font-size:18px;color:#6c27b0'>{line[4:]}</b>")
                elif line.startswith("## "): out.append(f"<b style='font-size:18px;color:#1565c0'>{line[3:]}</b>")
                elif line.startswith("# "): out.append(f"<b style='font-size:18px;color:#0d47a1'>{line[2:]}</b>")
                elif line.startswith("- ") or line.startswith("* "): out.append(f"&nbsp;&nbsp;• {line[2:]}")
                elif line.strip() == "": out.append("<br>")
                else: out.append(line)
            result = "<br>".join(out)
            result = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", result)
            result = re.sub(r"\*(.+?)\*", r"<i>\1</i>", result)
            return result

    def _on_suggestion_clicked(self, question: str):
        """User clicked a suggested follow-up question — send it immediately."""
        self.chat_input.setText(question)
        self._send_chat()

    def _on_chat_error(self, msg: str):
        self._stop_thinking()
        self.chat_display.append(f"<b style='color:#f38ba8'>Error:</b> {msg}")
        self.btn_send.setEnabled(True)
        self.btn_save_note.setEnabled(True)

    def _save_chat_as_note(self):
        """Save the current chat display content as a note in the active notebook."""
        if not self._current_notebook_id:
            return
        plain = self.chat_display.toPlainText().strip()
        if not plain:
            return

        from PySide6.QtWidgets import QInputDialog
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M')
        name, ok = QInputDialog.getText(self, "Save as Source",
            "Enter a name for this source:", text="")
        if not ok:
            return
        title = f"Chat {timestamp} — {name.strip()}" if name.strip() else f"Chat {timestamp}"
        self.btn_save_note.setEnabled(False)
        self.btn_save_note.setText("📌 Saving…")

        w = SaveNoteWorker(self._current_notebook_id, title, plain)
        w.done.connect(self._on_note_saved)
        w.error.connect(self._on_note_save_error)
        self._start_worker(w)

    def _on_note_saved(self, note_id: str):
        self.btn_save_note.setText("📌 Note")
        self.btn_save_note.setEnabled(True)
        self.chat_display.append(
            "<i style='color:#a6e3a1'>✓ Saved as note.</i><br>"
        )
        # Mở Studio panel nếu chưa mở, rồi refresh notes
        if not self._studio_widget.isVisible():
            self._toggle_studio_panel()
        else:
            self._load_notes()

    def _on_note_save_error(self, msg: str):
        self.btn_save_note.setText("📌 Note")
        self.btn_save_note.setEnabled(True)
        self.chat_display.append(
            f"<i style='color:#f38ba8'>✗ Failed to save note: {msg}</i><br>"
        )

    def _auto_export_excel_if_table(self, text: str):
        """If response contains a markdown table, auto-export and show link in chat."""
        import re
        table_lines = [l for l in text.splitlines() if re.match(r"\s*\|", l)]
        if len(table_lines) < 2:
            return
        import openpyxl, os, tempfile
        from openpyxl.styles import Font, PatternFill, Alignment


        rows_data = []
        for line in table_lines:
            if re.match(r"\s*\|[-:\s|]+\|?\s*$", line):
                continue
            cells = [c.strip() for c in line.strip().strip("|").split("|")]
            rows_data.append(cells)
        if not rows_data:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2E75B6")
        wrap_align  = Alignment(wrap_text=True, vertical="top")

        for ci, val in enumerate(rows_data[0], 1):
            cell = ws.cell(row=1, column=ci, value=val)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = wrap_align
        for ri, row in enumerate(rows_data[1:], 2):
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val).alignment = wrap_align
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 28

        tmp_dir = os.path.join(tempfile.gettempdir(), "finder_excel")
        os.makedirs(tmp_dir, exist_ok=True)
        fname = f"table_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path  = os.path.join(tmp_dir, fname)
        wb.save(path)

        self.chat_display.append(
            f"<a href='xlsx://{path}' style='color:#89b4fa'>📊 {fname} — Click to open</a><br>"
        )

    def _export_last_answer_to_excel(self):
        """Parse last Mr Finder answer and export to Excel."""
        text = getattr(self, "_last_answer", "").strip()
        if not text:
            return
        import re, os
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from PySide6.QtWidgets import QFileDialog


        path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel", f"NbLM_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel (*.xlsx)"
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Answer"

        header_font  = Font(bold=True, color="FFFFFF")
        header_fill  = PatternFill("solid", fgColor="2E75B6")
        wrap_align   = Alignment(wrap_text=True, vertical="top")

        # ── Detect markdown table ──────────────────────────────────
        table_lines = [l for l in text.splitlines() if re.match(r"\s*\|", l)]
        if len(table_lines) >= 2:
            rows_data = []
            for line in table_lines:
                if re.match(r"\s*\|[-:\s|]+\|?\s*$", line):
                    continue  # skip separator row
                cells = [c.strip() for c in line.strip().strip("|").split("|")]
                rows_data.append(cells)
            if rows_data:
                # Header row
                for ci, val in enumerate(rows_data[0], 1):
                    cell = ws.cell(row=1, column=ci, value=val)
                    cell.font  = header_font
                    cell.fill  = header_fill
                    cell.alignment = wrap_align
                for ri, row in enumerate(rows_data[1:], 2):
                    for ci, val in enumerate(row, 1):
                        ws.cell(row=ri, column=ci, value=val).alignment = wrap_align
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = 30
                wb.save(path)
                os.startfile(path)
                self.chat_display.append(
                    f"<i style='color:#a6e3a1'>✓ Exported table to {os.path.basename(path)}</i><br>"
                )
                return

        # ── Detect numbered / bullet list ─────────────────────────
        items = []
        for line in text.splitlines():
            m = re.match(r"^\s*(?:\d+[\.\)]|[-•*])\s+(.+)", line)
            if m:
                items.append(m.group(1).strip())
        if items:
            ws.cell(row=1, column=1, value="Item").font = header_font
            ws.cell(row=1, column=1).fill = header_fill
            ws.cell(row=1, column=1).alignment = wrap_align
            for ri, item in enumerate(items, 2):
                ws.cell(row=ri, column=1, value=item).alignment = wrap_align
            ws.column_dimensions["A"].width = 80
            wb.save(path)
            os.startfile(path)
            self.chat_display.append(
                f"<i style='color:#a6e3a1'>✓ Exported {len(items)} items to {os.path.basename(path)}</i><br>"
            )
            return

        # ── Fallback: plain paragraphs, 1 per row ─────────────────
        paras = [p.strip() for p in re.split(r"\n{2,}", text) if p.strip()]
        ws.cell(row=1, column=1, value="Content").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = wrap_align
        for ri, para in enumerate(paras, 2):
            ws.cell(row=ri, column=1, value=para).alignment = wrap_align
        ws.column_dimensions["A"].width = 100
        wb.save(path)
        os.startfile(path)
        self.chat_display.append(
            f"<i style='color:#a6e3a1'>✓ Exported to {os.path.basename(path)}</i><br>"
        )

    # ── Studio ────────────────────────────────────────────────────

    def _on_right_tab_changed(self, _index: int):
        pass  # reserved for future use

    def _toggle_studio_panel(self):
        visible = self._studio_widget.isVisible()
        if visible:
            self._studio_widget.setVisible(False)
            self.btn_toggle_studio.setChecked(False)
        else:
            self._studio_widget.setVisible(True)
            self.btn_toggle_studio.setChecked(True)
            # Set splitter sizes: 60% chat, 40% studio
            total = self._inner_splitter.width()
            self._inner_splitter.setSizes([int(total * 0.6), int(total * 0.4)])
            if self._current_notebook_id:
                self._load_studio()

    def _set_studio_btns_enabled(self, enabled: bool):
        for btn in self._studio_gen_btns.values():
            btn.setEnabled(enabled)

    def _load_studio(self):
        """Refresh both notes and artifacts lists."""
        self._load_notes()
        self._load_artifacts()

    def _load_notes(self):
        if not self._current_notebook_id:
            return
        self.lst_notes.clear()
        self.lst_notes.addItem("⏳ Loading…")
        w = ListNotesWorker(self._current_notebook_id)
        w.done.connect(self._on_notes_loaded)
        w.error.connect(lambda e: (
            self.lst_notes.clear(),
            self.lst_notes.addItem(f"Error: {e}")
        ))
        self._start_worker(w)

    def _on_notes_loaded(self, notes: list):
        self.lst_notes.clear()
        if not notes:
            self.lst_notes.addItem("(no notes yet)")
            return
        for n in notes:
            title   = getattr(n, "title", None) or "Untitled"
            nid     = getattr(n, "id", None) or ""
            content = getattr(n, "content", "") or ""
            item    = QListWidgetItem(f"📝 {title}")
            item.setData(Qt.UserRole,     nid)
            item.setData(Qt.UserRole + 1, content)
            item.setData(Qt.UserRole + 2, title)
            item.setToolTip(title)
            self.lst_notes.addItem(item)

    def _show_note_context_menu(self, pos):
        item = self.lst_notes.itemAt(pos)
        if not item or not item.data(Qt.UserRole):
            return
        from PySide6.QtWidgets import QMenu
        menu = QMenu(self)
        menu.addAction("👁 View").triggered.connect(
            lambda: self._open_note_item(item))
        menu.addAction("📎 Convert to Source").triggered.connect(
            lambda: self._convert_note_to_source(item))
        menu.addSeparator()
        menu.addAction("🗑 Delete").triggered.connect(
            lambda: self._delete_note(item))
        menu.exec(self.lst_notes.mapToGlobal(pos))

    def _open_note_item(self, item):
        title   = item.data(Qt.UserRole + 2) or item.text().lstrip("📝 ")
        content = item.data(Qt.UserRole + 1) or ""
        nid     = item.data(Qt.UserRole) or ""
        if not content:
            # content might not be loaded — fetch it
            self._fetch_and_show_note(nid, title)
            return
        self._show_note_panel(nid, title, content)

    def _fetch_and_show_note(self, note_id: str, title: str):
        class _FetchNote(QThread):
            done  = Signal(str, str)
            error = Signal(str)
            def __init__(self, nb_id, nid):
                super().__init__()
                self.nb_id = nb_id
                self.nid   = nid
            def run(self):
                try:
                    from notebooklm import NotebookLMClient
                    async def _g():
                        async with await NotebookLMClient.from_storage() as client:
                            n = await client.notes.get(self.nb_id, self.nid)
                            return getattr(n, "content", "") or ""
                    self.done.emit(self.nid, _run_async(_g()))
                except Exception as e:
                    self.error.emit(str(e))

        fw = _FetchNote(self._current_notebook_id, note_id)
        fw.done.connect(lambda nid, c: self._show_note_panel(nid, title, c))
        fw.error.connect(lambda e: self.lbl_studio_status.setText(f"❌ {e}"))
        self._start_worker(fw)

    def _show_note_panel(self, note_id: str, title: str, content: str):
        """Show note content in a dialog with Convert to Source button."""
        from PySide6.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout,
                                        QTextEdit, QPushButton, QLabel)
        dlg = QDialog(self)
        dlg.setWindowTitle(f"📝 {title}")
        dlg.resize(680, 500)
        lay = QVBoxLayout(dlg)

        lbl = QLabel(f"<b>{title}</b>")
        lbl.setStyleSheet("font-size:14px; padding:4px 0;")
        lay.addWidget(lbl)

        lbl_info = QLabel("(Saved responses are view only)")
        lbl_info.setStyleSheet("color:#888; font-size:12px;")
        lay.addWidget(lbl_info)

        txt = QTextEdit()
        txt.setReadOnly(True)
        txt.setPlainText(content)
        txt.setStyleSheet("font-size:14px;")
        lay.addWidget(txt, 1)

        btn_row = QHBoxLayout()
        btn_convert = QPushButton("📎 Convert to source")
        btn_convert.setFixedHeight(34)
        btn_delete  = QPushButton("🗑 Delete")
        btn_delete.setFixedHeight(34)
        btn_close   = QPushButton("Close")
        btn_close.setFixedHeight(34)
        btn_row.addWidget(btn_convert)
        btn_row.addWidget(btn_delete)
        btn_row.addStretch()
        btn_row.addWidget(btn_close)
        lay.addLayout(btn_row)

        def _do_convert():
            btn_convert.setEnabled(False)
            btn_convert.setText("Converting…")
            w = ConvertNoteToSourceWorker(
                self._current_notebook_id, note_id, title, content)
            w.done.connect(lambda sid: (
                btn_convert.setText("✅ Converted"),
                self._load_sources(),
            ))
            w.error.connect(lambda e: (
                btn_convert.setEnabled(True),
                btn_convert.setText("📎 Convert to source"),
                self.lbl_studio_status.setText(f"❌ {e}"),
            ))
            self._start_worker(w)

        def _do_delete():
            self._delete_note_by_id(note_id)
            dlg.accept()

        btn_convert.clicked.connect(_do_convert)
        btn_delete.clicked.connect(_do_delete)
        btn_close.clicked.connect(dlg.reject)
        dlg.exec()

    def _convert_note_to_source(self, item):
        nid     = item.data(Qt.UserRole)
        title   = item.data(Qt.UserRole + 2) or item.text().lstrip("📝 ")
        content = item.data(Qt.UserRole + 1) or ""
        if not content:
            self._fetch_and_show_note(nid, title)
            return
        w = ConvertNoteToSourceWorker(self._current_notebook_id, nid, title, content)
        w.done.connect(lambda _: (
            self.lbl_studio_status.setText("✅ Converted to source"),
            self._load_sources(),
        ))
        w.error.connect(lambda e: self.lbl_studio_status.setText(f"❌ {e}"))
        self._start_worker(w)
        self.lbl_studio_status.setText("⏳ Converting to source…")

    def _delete_note(self, item):
        from PySide6.QtWidgets import QMessageBox
        nid   = item.data(Qt.UserRole)
        reply = QMessageBox.question(self, "Delete Note",
            f'Delete "{item.text().strip()}"?',
            QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            self._delete_note_by_id(nid)

    def _delete_note_by_id(self, note_id: str):
        class _DelNote(QThread):
            done  = Signal()
            error = Signal(str)
            def __init__(self, nb_id, nid):
                super().__init__()
                self.nb_id = nb_id
                self.nid   = nid
            def run(self):
                try:
                    from notebooklm import NotebookLMClient
                    async def _d():
                        async with await NotebookLMClient.from_storage() as client:
                            await client.notes.delete(self.nb_id, self.nid)
                    _run_async(_d())
                    self.done.emit()
                except Exception as e:
                    self.error.emit(str(e))

        dw = _DelNote(self._current_notebook_id, note_id)
        dw.done.connect(self._load_notes)
        dw.error.connect(lambda e: self.lbl_studio_status.setText(f"❌ {e}"))
        self._start_worker(dw)

    def _load_artifacts(self):
        if not self._current_notebook_id:
            return
        self.lst_artifacts.clear()
        self.lst_artifacts.addItem("⏳ Loading…")
        w = ListArtifactsWorker(self._current_notebook_id)
        w.done.connect(self._on_artifacts_loaded)
        w.error.connect(lambda e: (
            self.lst_artifacts.clear(),
            self.lst_artifacts.addItem(f"Error: {e}")
        ))
        self._start_worker(w)

    # Numeric artifact type codes returned by NbLM API
    _ARTIFACT_TYPE_NUM = {
        "1": "audio", "2": "report", "3": "video",
        "4": "quiz",  "5": "mind_map", "6": "flashcards",
        "7": "infographic", "8": "slide_deck", "9": "data_table",
    }

    _ARTIFACT_ICONS = {
        "audio":       "🎙",
        "video":       "🎥",
        "report":      "📄",
        "briefing_doc":"📄",
        "study_guide": "📚",
        "blog_post":   "📝",
        "quiz":        "❓",
        "flashcards":  "🃏",
        "mind_map":    "🗺",
        "infographic": "🖼",
        "slide_deck":  "🎞",
        "data_table":  "📊",
        "unknown":     "📁",
    }

    def _parse_artifact_kind(self, a) -> str:
        raw = getattr(a, "artifact_type", None) or getattr(a, "type", None)
        if raw is None:
            return "unknown"
        # Enum with .value attribute
        if hasattr(raw, "value"):
            return str(raw.value)
        s = str(raw).replace("ArtifactType.", "").strip()
        # Numeric fallback
        return self._ARTIFACT_TYPE_NUM.get(s, s)

    def _on_artifacts_loaded(self, artifacts: list):
        self.lst_artifacts.clear()
        if not artifacts:
            self.lst_artifacts.addItem("(no artifacts yet)")
            return
        pending = getattr(self, "_pending_open_artifact_id", None)
        self._pending_open_artifact_id = None
        for a in artifacts:
            kind  = self._parse_artifact_kind(a)
            title = getattr(a, "title", None) or kind
            aid   = (getattr(a, "id", None) or getattr(a, "artifact_id", None)
                     or getattr(a, "task_id", None) or "")
            icon  = self._ARTIFACT_ICONS.get(kind, "📁")
            item  = QListWidgetItem(f"{icon} {title}")
            item.setData(Qt.UserRole,     aid)
            item.setData(Qt.UserRole + 1, kind)
            item.setToolTip(title)
            self.lst_artifacts.addItem(item)
            if pending and aid == pending:
                self._open_artifact_item(item)

    def _show_artifact_context_menu(self, pos):
        item = self.lst_artifacts.itemAt(pos)
        if not item or not item.data(Qt.UserRole):
            return
        from PySide6.QtWidgets import QMenu
        menu = QMenu(self)
        menu.addAction("📂 Open").triggered.connect(
            lambda: self._open_artifact_item(item))
        menu.addAction("🗑 Delete").triggered.connect(
            lambda: self._delete_artifact(item))
        menu.exec(self.lst_artifacts.mapToGlobal(pos))

    def _open_artifact_item(self, item):
        aid  = item.data(Qt.UserRole)
        kind = item.data(Qt.UserRole + 1) or ""
        if not aid or not self._current_notebook_id:
            return
        self.lbl_studio_status.setText(f"⏳ Opening {item.text().strip()}…")
        w = OpenArtifactWorker(self._current_notebook_id, aid, kind)
        w.done.connect(self._on_artifact_opened)

        def _on_open_error(msg: str, _kind=kind):
            self.lbl_studio_status.setText(f"❌ {msg}")
            if _kind == "infographic":
                QMessageBox.warning(
                    self, "Infographic",
                    "Could not download the infographic image.\n\n"
                    "The infographic may still be processing on Google's servers.\n"
                    "Please wait a moment and try again, or view it at notebooklm.google.com."
                )

        w.error.connect(_on_open_error)
        self._start_worker(w)

    def _on_artifact_opened(self, kind: str, data: str):
        import os
        self.lbl_studio_status.setText("")
        if kind == "file_path":
            os.startfile(data)
        elif kind == "image_path":
            os.startfile(data)
        elif kind == "csv_path":
            self._csv_to_excel_and_open(data)
        elif kind == "report_md":
            self._show_text_dialog("📄 Report", data, markdown=True)
        elif kind == "html_view":
            self._show_html_dialog(data)
        elif kind == "mind_map_saved":
            # data = JSON string of the saved mind map tree → render trực tiếp, không gọi lại API
            import json as _json
            try:
                tree = _json.loads(data)
            except Exception:
                tree = {"name": "Mind Map", "children": []}
            title = tree.get("name", "Mind Map")
            html = _mindmap_to_html(tree, title)
            # Dùng pdf_preview để hiển thị giống mind map local
            from ui.notebooklm_window import _mindmap_to_html as _mm2html
            _parent = self.window()
            pp = getattr(_parent, "pdf_preview", None)
            if pp and getattr(pp, "_web_view", None):
                if not pp.isVisible():
                    pp.show()
                    try:
                        _parent._splitter.setSizes([6000, 4000])
                    except Exception:
                        pass
                pp._web_view.setHtml(html)
                pp._show_mindmap_panel()
                pp._online_mindmap_html = html
            else:
                self._show_html_dialog(html)

    def _csv_to_excel_and_open(self, csv_path: str):
        import csv, os
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        xlsx_path = csv_path.replace(".csv", ".xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="2E75B6")
        wrap     = Alignment(wrap_text=True, vertical="top")
        with open(csv_path, encoding="utf-8", newline="") as f:
            for ri, row in enumerate(csv.reader(f), 1):
                for ci, val in enumerate(row, 1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.alignment = wrap
                    if ri == 1:
                        cell.font = hdr_font
                        cell.fill = hdr_fill
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 28
        wb.save(xlsx_path)
        os.startfile(xlsx_path)

    def _show_text_dialog(self, title: str, text: str, markdown: bool = False):
        from PySide6.QtWidgets import QDialog, QVBoxLayout, QTextEdit, QDialogButtonBox
        dlg = QDialog(self)
        dlg.setWindowTitle(title)
        dlg.resize(750, 550)
        lay = QVBoxLayout(dlg)
        txt = QTextEdit()
        txt.setReadOnly(True)
        if markdown:
            txt.setHtml(self._md_to_html(text))
        else:
            txt.setPlainText(text)
        lay.addWidget(txt)
        btns = QDialogButtonBox(QDialogButtonBox.Close)
        btns.rejected.connect(dlg.reject)
        lay.addWidget(btns)
        dlg.exec()

    def _show_html_dialog(self, html: str):
        from PySide6.QtWidgets import QDialog, QVBoxLayout, QDialogButtonBox
        try:
            from PySide6.QtWebEngineWidgets import QWebEngineView
            dlg = QDialog(self)
            dlg.setWindowTitle("Studio")
            dlg.resize(800, 600)
            lay = QVBoxLayout(dlg)
            web = QWebEngineView()
            web.setHtml(html)
            lay.addWidget(web)
            btns = QDialogButtonBox(QDialogButtonBox.Close)
            btns.rejected.connect(dlg.reject)
            lay.addWidget(btns)
            dlg.exec()
        except ImportError:
            # Fallback: plain text
            from PySide6.QtWidgets import QTextEdit, QDialog, QVBoxLayout, QDialogButtonBox
            import re
            plain = re.sub(r"<[^>]+>", "", html)
            self._show_text_dialog("Studio", plain)

    def _delete_artifact(self, item):
        from PySide6.QtWidgets import QMessageBox
        aid   = item.data(Qt.UserRole)
        title = item.text().strip()
        if not aid:
            QMessageBox.warning(self, "Error", "Artifact ID not found — cannot delete.")
            return
        reply = QMessageBox.question(self, "Delete Artifact",
            f'Delete "{title}"?', QMessageBox.Yes | QMessageBox.No)
        if reply != QMessageBox.Yes:
            return

        class _DelWorker(QThread):
            done  = Signal()
            error = Signal(str)
            def __init__(self, nb_id, aid):
                super().__init__()
                self.nb_id = nb_id
                self.aid   = aid
            def run(self):
                try:
                    from notebooklm import NotebookLMClient
                    async def _d():
                        async with await NotebookLMClient.from_storage() as client:
                            await client.artifacts.delete(self.nb_id, self.aid)
                    _run_async(_d())
                    self.done.emit()
                except Exception as e:
                    self.error.emit(str(e))
        dw = _DelWorker(self._current_notebook_id, aid)
        dw.done.connect(self._load_artifacts)
        dw.error.connect(lambda e: QMessageBox.critical(self, "Delete Error", e))
        self._start_worker(dw)

    def _studio_generate(self, kind: str):
        if not self._current_notebook_id:
            return
        from PySide6.QtWidgets import QInputDialog
        instructions = ""
        if kind not in ("mind_map", "audio"):
            instructions, ok = QInputDialog.getText(
                self, "Extra Instructions (optional)",
                "Thêm yêu cầu tùy chỉnh (hoặc bỏ trống):", text="")
            if not ok:
                return

        # Lấy source IDs đang được tick (None = dùng tất)
        checked = self._get_checked_source_ids()
        source_ids = checked if checked else None

        # Lưu lại để dùng khi mở mind map artifact
        if kind == "mind_map":
            self._pending_mindmap_source_ids = checked  # [] = all sources

        self._set_studio_btns_enabled(False)
        label = {
            "mind_map":    "🗺 Mind Map",
            "briefing_doc":"📄 Briefing Doc",
            "study_guide": "📚 Study Guide",
            "blog_post":   "📝 Blog Post",
            "flashcards":  "🃏 Flashcards",
            "quiz":        "❓ Quiz",
            "data_table":  "📊 Data Table",
            "infographic": "🖼 Infographic",
            "slide_deck":  "🎞 Slide Deck",
            "audio":       "🎙 Audio",
        }.get(kind, kind)
        src_info = f" ({len(source_ids)} sources)" if source_ids else ""
        self.lbl_studio_status.setText(f"⏳ Generating {label}{src_info}… (may take a while)")

        w = GenerateArtifactWorker(
            self._current_notebook_id, kind,
            lang=self._current_language,
            instructions=instructions,
            source_ids=source_ids,
        )
        w.done.connect(self._on_studio_generated)
        w.error.connect(self._on_studio_error)
        self._start_worker(w)

    def _on_studio_generated(self, artifact_id: str):
        self._set_studio_btns_enabled(True)
        self.lbl_studio_status.setText("✅ Done! Loading artifact…")
        self._load_artifacts()
        # Auto-open the new artifact after list refreshes
        self._pending_open_artifact_id = artifact_id

    def _on_studio_error(self, msg: str):
        self._set_studio_btns_enabled(True)
        self.lbl_studio_status.setText(f"❌ Error: {msg}")

    # ── Sources ───────────────────────────────────────────────────

    def _add_file(self):
        if not self._current_notebook_id:
            return
        path, _ = QFileDialog.getOpenFileName(
            self, "Select File", "",
            "Documents (*.pdf *.txt *.doc *.docx *.pptx *.md)"
        )
        if not path:
            return
        # Kiểm tra trùng với danh sách source đang hiển thị
        fname = os.path.basename(path).strip().lower()
        existing = []
        for i in range(self.lst_sources.count()):
            t = self.lst_sources.item(i).text().strip().lower()
            existing.append(t)
        if fname in existing or any(fname in t for t in existing):
            QMessageBox.warning(self, "Duplicate File",
                f'"{os.path.basename(path)}" already exists in this notebook.')
            return
        self.btn_add_file.setEnabled(False)
        self.chk_vision.setEnabled(False)
        use_vision = self.chk_vision.isChecked()
        w = AddSourceWorker(
            self._current_notebook_id, path,
            use_vision=use_vision,
        )
        w.done.connect(self._on_source_added)
        w.error.connect(lambda e: (
            QMessageBox.critical(self, "Error", e),
            self.btn_add_file.setEnabled(True),
            self.chk_vision.setEnabled(True),
        ))
        self._start_worker(w)

    def _on_source_added(self):
        self.btn_add_file.setEnabled(True)
        self.chk_vision.setEnabled(True)
        self._load_sources()  # reload danh sách thực tế

    # ── Folder batch upload ──────────────────────────────────────

    def _add_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Chọn folder để upload")
        if not folder:
            return

        SUPPORTED = BatchAddSourceWorker.SUPPORTED
        reply_sub = QMessageBox.question(
            self, "Bao gồm thư mục con?",
            "Scan cả các thư mục con bên trong không?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
        )
        recursive = (reply_sub == QMessageBox.Yes)

        files = []
        if recursive:
            for root, _dirs, fnames in os.walk(folder):
                for f in sorted(fnames):
                    if os.path.splitext(f)[1].lower() in SUPPORTED:
                        files.append(os.path.join(root, f))
        else:
            for f in sorted(os.listdir(folder)):
                fp = os.path.join(folder, f)
                if os.path.isfile(fp) and os.path.splitext(f)[1].lower() in SUPPORTED:
                    files.append(fp)

        if not files:
            QMessageBox.information(self, "Không tìm thấy file",
                "Không có file được hỗ trợ nào trong folder này.\n"
                f"Định dạng hỗ trợ: {', '.join(sorted(SUPPORTED))}")
            return

        folder_name = os.path.basename(folder)
        nb_count    = (len(files) - 1) // BatchAddSourceWorker.BATCH_SIZE + 1
        batch_size  = BatchAddSourceWorker.BATCH_SIZE

        lines = [f"Tìm thấy <b>{len(files)}</b> file trong '<b>{folder_name}</b>'.<br><br>"]
        lines.append(f"Sẽ upload lên <b>{nb_count}</b> notebook:<br>")
        for i in range(nb_count):
            start = i * batch_size + 1
            end   = min((i + 1) * batch_size, len(files))
            name  = folder_name if i == 0 else f"{folder_name} (Part {i + 1})"
            lines.append(f"  &nbsp;• <b>{name}</b> — {end - start + 1} file ({start}–{end})<br>")
        lines.append("<br>Bắt đầu upload?")

        confirm = QMessageBox(self)
        confirm.setWindowTitle("Xác nhận Batch Upload")
        confirm.setTextFormat(Qt.RichText)
        confirm.setText("".join(lines))
        confirm.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        confirm.setDefaultButton(QMessageBox.Yes)
        if confirm.exec() != QMessageBox.Yes:
            return

        self.btn_add_file.setEnabled(False)
        self.btn_add_folder.setEnabled(False)
        self.lbl_batch_progress.setVisible(True)
        self.lbl_batch_progress.setText("⏳ Chuẩn bị upload…")

        w = BatchAddSourceWorker(folder_name, files)
        w.progress.connect(self._on_batch_progress)
        w.notebook_created.connect(self._on_batch_notebook_created)
        w.file_error.connect(self._on_batch_file_error)
        w.relogin_status.connect(self._on_batch_relogin_status)
        w.done.connect(self._on_batch_done)
        self._start_worker(w)

    def _on_batch_progress(self, idx: int, total: int, fname: str, nb_title: str):
        self.lbl_batch_progress.setText(
            f"⬆ {idx}/{total} — <i>{fname}</i>  →  📓 {nb_title}"
        )

    def _on_batch_notebook_created(self, nb_id: str, nb_title: str):
        """Thêm notebook mới vừa tạo vào danh sách bên trái."""
        from PySide6.QtWidgets import QTreeWidgetItem
        item = QTreeWidgetItem([f"📓  {nb_title}"])
        item.setData(0, Qt.UserRole,     nb_id)
        item.setData(0, Qt.UserRole + 1, "notebook")
        font = item.font(0)
        font.setBold(True)
        item.setFont(0, font)
        from PySide6.QtWidgets import QTreeWidgetItem as _TWI
        placeholder = _TWI(["⏳"])
        placeholder.setData(0, Qt.UserRole + 1, "placeholder")
        item.addChild(placeholder)
        self.lst_notebooks.addTopLevelItem(item)

    def _on_batch_file_error(self, fname: str, msg: str):
        self.lbl_batch_progress.setText(f"⚠ Lỗi: {fname} — {msg[:80]}")

    def _on_batch_relogin_status(self, msg: str):
        self.lbl_batch_progress.setText(msg)

    def _on_batch_done(self, success: int, fail: int, skipped: int):
        self.btn_add_file.setEnabled(True)
        self.btn_add_folder.setEnabled(True)
        self.lbl_batch_progress.setVisible(False)
        self._reload_current_notebook_sources()
        result = f"✅ Upload xong: {success} thành công"
        if fail:
            result += f", {fail} lỗi"
        if skipped:
            result += f"\n⏭ {skipped} file bỏ qua (đã upload trước đó)"
        QMessageBox.information(self, "Batch Upload hoàn tất", result)

    # ── Merge notebooks ──────────────────────────────────────────

    def _merge_notebooks(self):
        from PySide6.QtWidgets import QInputDialog
        nb_titles = [t.split("  ", 1)[-1] for t in self._checked_notebook_ids.values()]
        default_name = " + ".join(nb_titles[:3]) + (" + …" if len(nb_titles) > 3 else "")
        new_title, ok = QInputDialog.getText(
            self, "Tên notebook mới", "Đặt tên cho notebook gộp:",
            text=default_name
        )
        if not ok or not new_title.strip():
            return

        self.btn_merge_nb.setEnabled(False)
        self.lbl_batch_progress.setVisible(True)
        self.lbl_batch_progress.setText("⏳ Đang chuẩn bị merge…")

        w = MergeNotebooksWorker(dict(self._checked_notebook_ids), new_title.strip())
        w.progress.connect(self.lbl_batch_progress.setText)
        w.notebook_created.connect(self._on_batch_notebook_created)
        w.done.connect(self._on_merge_done)
        w.error.connect(self._on_merge_error)
        self._start_worker(w)

    def _on_merge_done(self, n_files: int, n_parts: int):
        self.btn_merge_nb.setEnabled(len(self._checked_notebook_ids) >= 2)
        self.lbl_batch_progress.setVisible(False)
        self._load_notebooks()
        QMessageBox.information(
            self, "Merge hoàn tất",
            f"✅ Đã gộp {n_files} file thành {n_parts} phần vào notebook mới."
        )

    def _on_merge_error(self, msg: str):
        self.btn_merge_nb.setEnabled(len(self._checked_notebook_ids) >= 2)
        self.lbl_batch_progress.setVisible(False)
        QMessageBox.critical(self, "Merge thất bại", msg)

    # ── Source checkbox helpers ──────────────────────────────────

    def _on_source_check_changed(self, _item):
        self._update_src_action_bar()

    def _update_src_action_bar(self):
        checked = self._get_checked_source_items()
        has = bool(checked)
        total = self.lst_sources.count()
        if total == 0:
            self.chk_select_all_src.setText("☐ Select All")
        elif len(checked) == total:
            self.chk_select_all_src.setText("☑ Select All")
        else:
            self.chk_select_all_src.setText("☐ Select All")

    def _get_checked_source_ids(self) -> list[str]:
        """Trả về list source_id đang được tick — ưu tiên lấy từ tree."""
        ids = []
        # Tìm notebook node hiện tại trong tree
        for i in range(self.lst_notebooks.topLevelItemCount()):
            nb_item = self.lst_notebooks.topLevelItem(i)
            if nb_item.data(0, Qt.UserRole) == self._current_notebook_id:
                for j in range(nb_item.childCount()):
                    child = nb_item.child(j)
                    if (child.data(0, Qt.UserRole + 1) == "source"
                            and child.checkState(0) == Qt.Checked):
                        ids.append(child.data(0, Qt.UserRole))
                return ids
        # Fallback: lấy từ lst_sources
        for i in range(self.lst_sources.count()):
            it = self.lst_sources.item(i)
            if it and it.checkState() == Qt.Checked and it.data(Qt.UserRole):
                ids.append(it.data(Qt.UserRole))
        return ids

    def _get_checked_source_items(self):
        """Dùng cho delete selected — lấy từ lst_sources."""
        items = []
        for i in range(self.lst_sources.count()):
            it = self.lst_sources.item(i)
            if it and it.checkState() == Qt.Checked and it.data(Qt.UserRole):
                items.append(it)
        return items

    def _toggle_select_all_sources(self):
        checked = self._get_checked_source_items()
        total = self.lst_sources.count()
        # Nếu chưa chọn hết → chọn tất; nếu đã chọn hết → bỏ chọn tất
        new_state = Qt.Unchecked if len(checked) == total else Qt.Checked
        self.lst_sources.blockSignals(True)
        for i in range(total):
            it = self.lst_sources.item(i)
            if it and it.data(Qt.UserRole):
                it.setCheckState(new_state)
        self.lst_sources.blockSignals(False)
        self._update_src_action_bar()




# ── Dialog wrapper (backward compat) ─────────────────────────────

class NotebookLMWindow(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("NotebookLM")
        self.resize(1000, 680)
        self.setModal(False)
        self.setStyleSheet("QDialog { background: #1e2030; }")

        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.addWidget(NotebookLMWidget(self))

        from PySide6.QtWidgets import QApplication
        screen = QApplication.primaryScreen().availableGeometry()
        self.move(
            screen.center().x() - self.width() // 2,
            screen.center().y() - self.height() // 2,
        )
